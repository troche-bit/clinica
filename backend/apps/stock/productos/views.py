from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from django.db.models import Count, Q
from django.http import HttpResponse
from django.template.loader import render_to_string
from datetime import date
from collections import OrderedDict

from apps.administracion.auditoria.mixins import AuditoriaMixin
from apps.core.permissions import IsAdminRole, IsAdminOrRecepcionista
from .models import Grupo, ProductoServicio
from .serializers import (
    GrupoListSerializer, GrupoSerializer,
    ProductoServicioListSerializer, ProductoServicioSerializer,
)


class GrupoViewSet(AuditoriaMixin, viewsets.ModelViewSet):
    filter_backends    = [filters.SearchFilter, filters.OrderingFilter]
    search_fields      = ['descripcion']
    ordering_fields    = ['descripcion', 'activo']
    ordering           = ['descripcion']

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [IsAuthenticated()]
        if self.action in ('destroy', 'eliminados'):
            return [IsAuthenticated(), IsAdminRole()]
        return [IsAuthenticated(), IsAdminOrRecepcionista()]

    def get_serializer_class(self):
        if self.action in ['list', 'retrieve', 'eliminados']:
            return GrupoListSerializer
        return GrupoSerializer

    def get_queryset(self):
        qs = Grupo.objects.filter(is_deleted=False).annotate(
            total_productos=Count(
                'productos',
                filter=Q(productos__is_deleted=False, productos__activo=True),
            )
        )
        activo = self.request.query_params.get('activo')
        if activo == 'true':
            qs = qs.filter(activo=True)
        elif activo == 'false':
            qs = qs.filter(activo=False)
        return qs

    def perform_destroy(self, instance):
        if instance.productos.filter(is_deleted=False, activo=True).exists():
            raise ValidationError('No se puede eliminar un grupo con productos activos vinculados.')
        super().perform_destroy(instance)

    @action(detail=False, methods=['get'], url_path='eliminados')
    def eliminados(self, request):
        qs = Grupo.objects.filter(is_deleted=True).annotate(
            total_productos=Count('productos', filter=Q(productos__is_deleted=False))
        )
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)


class ProductoServicioViewSet(AuditoriaMixin, viewsets.ModelViewSet):
    filter_backends  = [filters.SearchFilter, filters.OrderingFilter]
    search_fields    = ['descripcion']
    ordering_fields  = ['descripcion', 'impuesto', 'activo']
    ordering         = ['descripcion']

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [IsAuthenticated()]
        if self.action in ('destroy', 'eliminados'):
            return [IsAuthenticated(), IsAdminRole()]
        return [IsAuthenticated(), IsAdminOrRecepcionista()]

    def get_serializer_class(self):
        if self.action in ['list', 'retrieve', 'eliminados']:
            return ProductoServicioListSerializer
        return ProductoServicioSerializer

    def get_queryset(self):
        qs = ProductoServicio.objects.filter(is_deleted=False).select_related('grupo')
        grupo_id = self.request.query_params.get('grupo')
        if grupo_id:
            qs = qs.filter(grupo_id=grupo_id)
        activo = self.request.query_params.get('activo')
        if activo == 'true':
            qs = qs.filter(activo=True)
        elif activo == 'false':
            qs = qs.filter(activo=False)
        return qs

    def perform_destroy(self, instance):
        from apps.facturacion.ventas.models import VentaFactDet
        if VentaFactDet.objects.filter(prs=instance, is_deleted=False).exists():
            raise ValidationError('No se puede eliminar: el producto tiene facturas emitidas vinculadas.')
        super().perform_destroy(instance)

    @action(detail=False, methods=['get'], url_path='eliminados')
    def eliminados(self, request):
        qs = ProductoServicio.objects.filter(is_deleted=True).select_related('grupo')
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    def _productos_agrupados(self, request):
        qs = (
            ProductoServicio.objects
            .filter(is_deleted=False, grupo__is_deleted=False)
            .select_related('grupo')
            .order_by('grupo__descripcion', 'descripcion')
        )
        grupo_id = request.query_params.get('grupo')
        if grupo_id:
            qs = qs.filter(grupo_id=grupo_id)

        agrupado = OrderedDict()
        for prod in qs:
            clave = prod.grupo.descripcion
            if clave not in agrupado:
                agrupado[clave] = []
            agrupado[clave].append({
                'descripcion': prod.descripcion,
                'impuesto':    prod.get_impuesto_display(),
                'activo':      prod.activo,
            })
        return [{'nombre': k, 'productos': v} for k, v in agrupado.items()]

    @action(detail=False, methods=['get'], url_path='reporte-productos')
    def reporte_productos(self, request):
        grupos = self._productos_agrupados(request)
        total  = sum(len(g['productos']) for g in grupos)
        hoy    = date.today()

        filtros_aplicados = []
        grupo_id = request.query_params.get('grupo')
        if grupo_id:
            try:
                grupo_obj = Grupo.objects.get(pk=grupo_id, is_deleted=False)
                filtros_aplicados.append(('Grupo', grupo_obj.descripcion))
            except Grupo.DoesNotExist:
                pass

        ctx  = {'fecha': hoy, 'total': total, 'grupos': grupos, 'filtros_aplicados': filtros_aplicados}
        html = render_to_string('informes/producto_lista.html', ctx, request=request)
        try:
            import weasyprint
            pdf = weasyprint.HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
        except Exception as e:
            return HttpResponse(f'Error generando PDF: {e}', status=500)
        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = 'inline; filename="productos.pdf"'
        return resp

    @action(detail=False, methods=['get'], url_path='reporte-productos-excel')
    def reporte_productos_excel(self, request):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse('openpyxl no está instalado.', status=500)

        grupos = self._productos_agrupados(request)
        total  = sum(len(g['productos']) for g in grupos)
        hoy    = date.today()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Productos por Grupo'

        fill_header = PatternFill('solid', fgColor='1A3A5C')
        fill_grupo  = PatternFill('solid', fgColor='EEF2F7')
        fill_par    = PatternFill('solid', fgColor='F8FAFC')
        font_header = Font(color='FFFFFF', bold=True, size=10)
        font_grupo  = Font(color='1A3A5C', bold=True, size=10)
        font_titulo = Font(color='1A3A5C', bold=True, size=13)
        font_meta   = Font(color='555555', size=9)
        thin_border = Border(bottom=Side(style='thin', color='E8EDF2'))
        align_c     = Alignment(horizontal='center', vertical='center')
        align_l     = Alignment(horizontal='left',   vertical='center')

        COLS = 4
        fila = 1

        ws.merge_cells(f'A{fila}:{get_column_letter(COLS)}{fila}')
        c = ws.cell(fila, 1, 'Clínica Lichi — Listado de Productos por Grupo')
        c.font = font_titulo; c.alignment = align_l
        ws.row_dimensions[fila].height = 20
        fila += 1

        ws.merge_cells(f'A{fila}:{get_column_letter(COLS)}{fila}')
        c = ws.cell(fila, 1, f"Generado el {hoy.strftime('%d/%m/%Y')}  —  {total} registro{'s' if total != 1 else ''}")
        c.font = font_meta; c.alignment = align_l
        fila += 2

        for col, txt in enumerate(['N°', 'Descripción', 'Impuesto', 'Activo'], 1):
            c = ws.cell(fila, col, txt)
            c.fill = fill_header; c.font = font_header
            c.alignment = align_c if col in (1, 4) else align_l
        ws.row_dimensions[fila].height = 18
        fila += 1

        nro_global = 0
        for grupo in grupos:
            ws.merge_cells(f'A{fila}:{get_column_letter(COLS)}{fila}')
            c = ws.cell(fila, 1, grupo['nombre'])
            c.fill = fill_grupo; c.font = font_grupo; c.alignment = align_l
            ws.row_dimensions[fila].height = 16
            fila += 1

            for nro, prod in enumerate(grupo['productos'], 1):
                nro_global += 1
                es_par = (nro % 2 == 0)
                vals   = [nro_global, prod['descripcion'], prod['impuesto'], 'Sí' if prod['activo'] else 'No']
                for col, val in enumerate(vals, 1):
                    c = ws.cell(fila, col, val)
                    if es_par: c.fill = fill_par
                    c.border    = thin_border
                    c.alignment = align_c if col in (1, 4) else align_l
                ws.row_dimensions[fila].height = 15
                fila += 1

        for i, ancho in enumerate([5, 42, 14, 8], 1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="productos_{hoy.strftime("%Y%m%d")}.xlsx"'
        return resp
