from datetime import timedelta, date as date_type
from decimal import Decimal

from django.db import models, transaction
from django.db.models import Max, Count, Sum, Q
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone

from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response

from apps.administracion.auditoria.mixins import AuditoriaMixin, _serializar
from apps.core.permissions import IsAdminRole, IsAdminOrRecepcionista
from apps.facturacion.configuracion.timbrado.models import Timbrado
from apps.stock.productos.models import ProductoServicio
from apps.forma_pago.models import FormaPago
from apps.finanzas.caja_banco.models import CuentaMcb, MovimientoCajaBanco
from apps.finanzas.cobranzas.models import CobranzaDet
from apps.administracion.persona.models import Persona
from apps.finanzas.estadocuenta.models import CtaCobrar

from .models import VentaFactCab, VentaFactDet, VentaFactDetCobranza
from .serializers import (
    VentaFactCabListSerializer,
    VentaFactCabDetalleSerializer,
    VentaFactCabUpdateSerializer,
    VentaFactCreateSerializer,
)
from apps.facturacion.services import calcular_item, calcular_totales


class VentaFactCabViewSet(AuditoriaMixin, viewsets.ModelViewSet):
    filter_backends    = [filters.SearchFilter, filters.OrderingFilter]
    search_fields      = ['persona__razon_social', 'persona__nro_documento', 'nro_comprobante']
    ordering_fields    = ['fecha', 'monto_total', 'fecha_creacion']
    ordering           = ['-fecha', '-fecha_creacion']

    def get_permissions(self):
        if self.action == 'pdf':
            return [AllowAny()]
        if self.action in ('list', 'retrieve', 'validar_timbrado', 'siguiente_numero',
                           'reporte_pdf', 'reporte_excel', 'reporte_control_pdf', 'reporte_control_excel',
                           'dashboard_mensual', 'estado_cuenta_pdf', 'estado_cuenta_excel',
                           'extracto_cuenta_pdf', 'extracto_cuenta_excel'):
            return [IsAuthenticated()]
        if self.action == 'destroy':
            return [IsAuthenticated(), IsAdminRole()]
        return [IsAuthenticated(), IsAdminOrRecepcionista()]



    def get_queryset(self):
        qs = VentaFactCab.objects.filter(is_deleted=False).select_related('persona', 'timbrado')

        condicion = self.request.query_params.get('condicion_vta')
        if condicion == 'true':
            qs = qs.filter(condicion_vta=True)
        elif condicion == 'false':
            qs = qs.filter(condicion_vta=False)

        fecha_desde = self.request.query_params.get('fecha_desde')
        fecha_hasta = self.request.query_params.get('fecha_hasta')
        if fecha_desde:
            qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha__lte=fecha_hasta)

        return qs

    def get_serializer_class(self):
        if self.action == 'create':
            return VentaFactCreateSerializer
        if self.action in ('update', 'partial_update'):
            return VentaFactCabUpdateSerializer
        if self.action == 'retrieve':
            return VentaFactCabDetalleSerializer
        return VentaFactCabListSerializer

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        ser = VentaFactCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        try:
            timbrado = Timbrado.objects.get(pk=data['timbrado'], is_deleted=False)
        except Timbrado.DoesNotExist:
            raise ValidationError({'timbrado': 'Timbrado no encontrado.'})

        hoy = timezone.localtime().date()
        if not (timbrado.inicio_vigencia <= hoy <= timbrado.fin_vigencia):
            raise ValidationError({'timbrado': 'El timbrado no está vigente.'})

        nro = data['nro_comprobante']
        if not (timbrado.nro_desde <= nro <= timbrado.nro_hasta):
            raise ValidationError({'nro_comprobante': f'Número fuera del rango del timbrado ({timbrado.nro_desde}–{timbrado.nro_hasta}).'})

        if VentaFactCab.objects.filter(timbrado=timbrado, nro_comprobante=nro, is_deleted=False).exists():
            raise ValidationError({'nro_comprobante': 'Este número de comprobante ya está en uso.'})

        items_data = []
        for det in data['detalle']:
            try:
                prod = ProductoServicio.objects.get(pk=det['prs'], is_deleted=False, activo=True)
            except ProductoServicio.DoesNotExist:
                raise ValidationError({'detalle': f'Producto ID {det["prs"]} no encontrado o inactivo.'})

            calcs = calcular_item(det['monto'], prod.impuesto)
            items_data.append({
                'prod': prod,
                'cantidad': det['cantidad'],
                'monto': det['monto'],
                'impuesto': prod.impuesto,
                'calcs': calcs,
            })

        totales = calcular_totales([it['calcs'] for it in items_data])

        vuelto = Decimal('0')
        if data['condicion_vta'] and data.get('cobranza'):
            total_cobrado = sum(Decimal(str(c['monto'])) for c in data['cobranza'])
            vuelto = max(Decimal('0'), total_cobrado - totales['monto_total'])

        try:
            persona = Persona.objects.get(pk=data['persona'], is_deleted=False)
        except Persona.DoesNotExist:
            raise ValidationError({'persona': 'Persona no encontrada.'})

        cab = VentaFactCab.objects.create(
            fecha           = data['fecha'],
            condicion_vta   = data['condicion_vta'],
            persona         = persona,
            timbrado        = timbrado,
            establecimiento = str(timbrado.punto_sucursal).zfill(3),
            expedicion      = str(timbrado.punto_expedicion).zfill(3),
            observacion     = data.get('observacion', ''),
            nro_comprobante = nro,
            vuelto          = vuelto,
            id_usu_creator  = request.user,
            **totales,
        )

        for it in items_data:
            VentaFactDet.objects.create(
                vfc            = cab,
                prs            = it['prod'],
                cantidad       = it['cantidad'],
                monto          = it['monto'],
                impuesto       = it['impuesto'],
                id_usu_creator = request.user,
                **it['calcs'],
            )

        if data['condicion_vta']:
            monto_pendiente = totales['monto_total']
            for cobr in data.get('cobranza', []):
                try:
                    fp  = FormaPago.objects.get(pk=cobr['forma_pago'])
                    cta = CuentaMcb.objects.get(pk=cobr['cta'], is_deleted=False)
                except (FormaPago.DoesNotExist, CuentaMcb.DoesNotExist) as e:
                    raise ValidationError({'cobranza': str(e)})

                det_cobr = VentaFactDetCobranza.objects.create(
                    vfc             = cab,
                    forma_pago      = fp,
                    cta             = cta,
                    monto           = cobr['monto'],
                    voucher         = cobr.get('voucher', ''),
                    nro_comprobante = cobr.get('nro_comprobante', ''),
                    id_usu_creator  = request.user,
                )

                monto_mov = min(Decimal(str(cobr['monto'])), monto_pendiente)
                if monto_mov > 0:
                    MovimientoCajaBanco.objects.create(
                        cta             = cta,
                        fecha           = data['fecha'],
                        nro_comprobante = cobr.get('nro_comprobante', '') or None,
                        monto_ingreso   = monto_mov,
                        monto_egreso    = Decimal('0'),
                        vuelto          = Decimal('0'),
                        vfdc_id         = det_cobr.id,
                        id_usu_creator  = request.user,
                    )
                    monto_pendiente -= monto_mov
        else:
            cfg  = data['cuotas']
            cant = cfg['cant_cuota']
            dias = cfg['dias_entre_cuotas']
            mt   = totales['monto_total']
            mc   = (mt / cant).quantize(Decimal('0.01'))

            for i in range(1, cant + 1):
                fecha_venc = data['fecha'] + timedelta(days=dias * i)
                CtaCobrar.objects.create(
                    vfc               = cab,
                    nro_cuota         = i,
                    cant_cuota        = cant,
                    monto_total       = mt,
                    monto_cuota       = mc,
                    saldo             = mc,
                    fecha_vencimiento = fecha_venc,
                    estado            = 'pendiente',
                    id_usu_creator    = request.user,
                )

        self._registrar('CREAR', cab, datos_antes=None, datos_despues=_serializar(cab))
        out = VentaFactCabDetalleSerializer(cab)
        return Response(out.data, status=201)

    @transaction.atomic
    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()

        if instance.is_anulado:
            raise ValidationError({'detail': 'No se puede editar una factura anulada.'})

        if 'detalle' in request.data:
            return self._full_update(request, instance)

        ser = VentaFactCabUpdateSerializer(instance, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        self.perform_update(ser)
        instance.refresh_from_db()
        return Response(VentaFactCabDetalleSerializer(instance).data)

    def _full_update(self, request, instance):
        # Permite cambiar condicion_vta en edición; si no viene en el request, mantiene la original
        raw_cond = request.data.get('condicion_vta', instance.condicion_vta)
        if isinstance(raw_cond, str):
            nueva_condicion = raw_cond.lower() in ('true', '1')
        else:
            nueva_condicion = bool(raw_cond)

        # Si la condición original era crédito, bloquear si hay cobros registrados en el módulo de cobranzas
        if not instance.condicion_vta:
            cobros_qs = CobranzaDet.objects.filter(
                cta_cobrar__vfc=instance, is_deleted=False, cobranza__is_deleted=False
            )
            if cobros_qs.exists():
                nros = list(cobros_qs.values_list('cobranza__comprobante_nro', flat=True).distinct())
                nros_str = ', '.join(str(n) for n in nros if n is not None)
                raise ValidationError({
                    'detail': f'No se puede editar: el comprobante tiene cobros registrados (Cobranza N° {nros_str}). Elimine la cobranza para poder modificarlo.'
                })

            tiene_pagos = (
                instance.cuotas.filter(is_deleted=False, estado='pagado').exists()
                or instance.cuotas.filter(is_deleted=False).exclude(saldo=models.F('monto_cuota')).exists()
            )
            if tiene_pagos:
                raise ValidationError({'detail': 'No se puede editar: ya existen cuotas cobradas en este comprobante.'})

        # Determinar establecimiento, expedición y timbrado (pueden cambiar en edición)
        estab      = request.data.get('establecimiento', instance.establecimiento)
        expedicion = request.data.get('expedicion',     instance.expedicion)

        if estab != instance.establecimiento or expedicion != instance.expedicion:
            hoy = timezone.localtime().date()
            timbrado = Timbrado.objects.filter(
                punto_sucursal=estab,
                punto_expedicion=expedicion,
                inicio_vigencia__lte=hoy,
                fin_vigencia__gte=hoy,
                is_deleted=False,
            ).order_by('-inicio_vigencia').first()
            if not timbrado:
                raise ValidationError({'establecimiento': f'No hay timbrado vigente para {estab}-{expedicion}.'})
        else:
            timbrado = instance.timbrado

        # Validar nro_comprobante
        raw_nro = request.data.get('nro_comprobante')
        if raw_nro is not None:
            try:
                nuevo_nro = int(raw_nro)
            except (ValueError, TypeError):
                raise ValidationError({'nro_comprobante': 'Número de comprobante inválido.'})
        else:
            nuevo_nro = instance.nro_comprobante

        if nuevo_nro != instance.nro_comprobante or timbrado.id != instance.timbrado_id:
            if not (timbrado.nro_desde <= nuevo_nro <= timbrado.nro_hasta):
                raise ValidationError({
                    'nro_comprobante': f'Número fuera del rango del timbrado ({timbrado.nro_desde}–{timbrado.nro_hasta}).'
                })
            if VentaFactCab.objects.filter(
                timbrado=timbrado, nro_comprobante=nuevo_nro, is_deleted=False
            ).exclude(pk=instance.pk).exists():
                raise ValidationError({'nro_comprobante': 'Este número de comprobante ya está en uso.'})

        ser = VentaFactCreateSerializer(data={
            'fecha':           request.data.get('fecha'),
            'condicion_vta':   nueva_condicion,
            'persona':         request.data.get('persona'),
            'timbrado':        timbrado.id,
            'observacion':     request.data.get('observacion', ''),
            'nro_comprobante': nuevo_nro,
            'detalle':         request.data.get('detalle', []),
            'cobranza':        request.data.get('cobranza', []),
            'cuotas':          request.data.get('cuotas'),
        })
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        now  = timezone.now()
        user = request.user

        cobr_ids = list(instance.cobranza.filter(is_deleted=False).values_list('id', flat=True))
        MovimientoCajaBanco.objects.filter(vfdc_id__in=cobr_ids, is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=user
        )
        instance.detalle.filter(is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=user
        )
        instance.cobranza.filter(is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=user
        )
        instance.cuotas.filter(is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=user
        )

        items_data = []
        for det in data['detalle']:
            try:
                prod = ProductoServicio.objects.get(pk=det['prs'], is_deleted=False, activo=True)
            except ProductoServicio.DoesNotExist:
                raise ValidationError({'detalle': f'Producto ID {det["prs"]} no encontrado o inactivo.'})
            calcs = calcular_item(det['monto'], prod.impuesto)
            items_data.append({
                'prod': prod, 'cantidad': det['cantidad'], 'monto': det['monto'],
                'impuesto': prod.impuesto, 'calcs': calcs,
            })

        totales = calcular_totales([it['calcs'] for it in items_data])

        vuelto = Decimal('0')
        if nueva_condicion and data.get('cobranza'):
            total_cobrado = sum(Decimal(str(c['monto'])) for c in data['cobranza'])
            vuelto = max(Decimal('0'), total_cobrado - totales['monto_total'])

        try:
            persona = Persona.objects.get(pk=data['persona'], is_deleted=False)
        except Persona.DoesNotExist:
            raise ValidationError({'persona': 'Persona no encontrada.'})

        instance.fecha           = data['fecha']
        instance.persona         = persona
        instance.observacion     = data.get('observacion', '')
        instance.condicion_vta   = nueva_condicion
        instance.nro_comprobante = nuevo_nro
        instance.establecimiento = estab
        instance.expedicion      = expedicion
        instance.timbrado        = timbrado
        instance.vuelto          = vuelto
        for k, v in totales.items():
            setattr(instance, k, v)
        instance.id_usu_modificator = user
        instance.save()

        for it in items_data:
            VentaFactDet.objects.create(
                vfc=instance, prs=it['prod'], cantidad=it['cantidad'],
                monto=it['monto'], impuesto=it['impuesto'],
                id_usu_creator=user, **it['calcs'],
            )

        if nueva_condicion:
            monto_pendiente = totales['monto_total']
            for cobr in data.get('cobranza', []):
                try:
                    fp  = FormaPago.objects.get(pk=cobr['forma_pago'])
                    cta = CuentaMcb.objects.get(pk=cobr['cta'], is_deleted=False)
                except (FormaPago.DoesNotExist, CuentaMcb.DoesNotExist) as e:
                    raise ValidationError({'cobranza': str(e)})
                det_cobr = VentaFactDetCobranza.objects.create(
                    vfc=instance, forma_pago=fp, cta=cta,
                    monto=cobr['monto'], voucher=cobr.get('voucher', ''),
                    nro_comprobante=cobr.get('nro_comprobante', ''),
                    id_usu_creator=user,
                )
                monto_mov = min(Decimal(str(cobr['monto'])), monto_pendiente)
                if monto_mov > 0:
                    MovimientoCajaBanco.objects.create(
                        cta=cta, fecha=data['fecha'],
                        nro_comprobante=cobr.get('nro_comprobante', '') or None,
                        monto_ingreso=monto_mov, monto_egreso=Decimal('0'),
                        vuelto=Decimal('0'), vfdc_id=det_cobr.id,
                        id_usu_creator=user,
                    )
                    monto_pendiente -= monto_mov
        else:
            cfg  = data['cuotas']
            cant = cfg['cant_cuota']
            dias = cfg['dias_entre_cuotas']
            mt   = totales['monto_total']
            mc   = (mt / cant).quantize(Decimal('0.01'))
            for i in range(1, cant + 1):
                fecha_venc = data['fecha'] + timedelta(days=dias * i)
                CtaCobrar.objects.create(
                    vfc=instance, nro_cuota=i, cant_cuota=cant,
                    monto_total=mt, monto_cuota=mc, saldo=mc,
                    fecha_vencimiento=fecha_venc, estado='pendiente',
                    id_usu_creator=user,
                )

        instance.refresh_from_db()
        return Response(VentaFactCabDetalleSerializer(instance).data)

    def perform_destroy(self, instance):
        if not instance.condicion_vta:
            cobros_qs = CobranzaDet.objects.filter(
                cta_cobrar__vfc=instance, is_deleted=False, cobranza__is_deleted=False
            )
            if cobros_qs.exists():
                nros = list(cobros_qs.values_list('cobranza__comprobante_nro', flat=True).distinct())
                nros_str = ', '.join(str(n) for n in nros if n is not None)
                raise ValidationError({
                    'detail': f'No se puede eliminar: el comprobante tiene cobros registrados (Cobranza N° {nros_str}). Elimine la cobranza primero.'
                })

        now      = timezone.now()
        user     = self.request.user
        cobr_ids = list(instance.cobranza.filter(is_deleted=False).values_list('id', flat=True))

        MovimientoCajaBanco.objects.filter(vfdc_id__in=cobr_ids, is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=user
        )
        instance.detalle.filter(is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=user
        )
        instance.cobranza.filter(is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=user
        )
        instance.cuotas.filter(is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=user
        )
        super().perform_destroy(instance)

    @action(detail=True, methods=['post'], url_path='anular')
    def anular(self, request, pk=None):
        instance = self.get_object()

        if instance.is_anulado:
            raise ValidationError({'detail': 'La factura ya está anulada.'})

        cobros_qs = CobranzaDet.objects.filter(
            cta_cobrar__vfc=instance, is_deleted=False, cobranza__is_deleted=False
        )
        if cobros_qs.exists():
            nros = list(cobros_qs.values_list('cobranza__comprobante_nro', flat=True).distinct())
            nros_str = ', '.join(str(n) for n in nros if n is not None)
            raise ValidationError({
                'detail': f'No se puede anular: tiene cobros registrados (Cobranza N° {nros_str}). Elimine la cobranza primero.'
            })

        now      = timezone.now()
        user     = request.user
        cobr_ids = list(instance.cobranza.filter(is_deleted=False).values_list('id', flat=True))

        MovimientoCajaBanco.objects.filter(vfdc_id__in=cobr_ids, is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=user
        )
        instance.detalle.filter(is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=user
        )
        instance.cobranza.filter(is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=user
        )
        instance.cuotas.filter(is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=user
        )

        instance.is_anulado        = True
        instance.id_usu_modificator = user
        instance.save(update_fields=['is_anulado', 'id_usu_modificator'])

        return Response(VentaFactCabDetalleSerializer(instance).data)

    @action(detail=False, methods=['post'], url_path='validar-timbrado')
    def validar_timbrado(self, request):
        estab      = request.data.get('establecimiento', '').strip()
        expedicion = request.data.get('expedicion', '').strip()
        numero     = request.data.get('numero')

        if not all([estab, expedicion, numero is not None]):
            return Response({'valido': False, 'mensaje': 'Datos incompletos.'}, status=400)

        hoy = timezone.localtime().date()
        try:
            numero = int(numero)
        except (ValueError, TypeError):
            return Response({'valido': False, 'mensaje': 'Número inválido.'}, status=400)

        qs = Timbrado.objects.filter(
            punto_sucursal=estab,
            punto_expedicion=expedicion,
            inicio_vigencia__lte=hoy,
            fin_vigencia__gte=hoy,
            is_deleted=False,
        )
        if not qs.exists():
            return Response({
                'valido': False,
                'mensaje': f'No hay timbrado vigente para {estab}-{expedicion}.',
                'timbrado_id': None,
            })

        timbrado = qs.order_by('-inicio_vigencia').first()

        if not (timbrado.nro_desde <= numero <= timbrado.nro_hasta):
            return Response({
                'valido': False,
                'mensaje': f'Número fuera del rango habilitado ({timbrado.nro_desde}–{timbrado.nro_hasta}).',
                'timbrado_id': timbrado.id,
            })

        if VentaFactCab.objects.filter(timbrado=timbrado, nro_comprobante=numero, is_deleted=False).exists():
            return Response({
                'valido': False,
                'mensaje': 'Este número ya fue utilizado.',
                'timbrado_id': timbrado.id,
            })

        return Response({
            'valido':       True,
            'timbrado_id':  timbrado.id,
            'nro_timbrado': timbrado.nro_timbrado,
            'mensaje':      f'Timbrado {timbrado.nro_timbrado} válido.',
        })

    @action(detail=False, methods=['get'], url_path='siguiente-numero')
    def siguiente_numero(self, request):
        estab      = request.query_params.get('establecimiento', '').strip()
        expedicion = request.query_params.get('expedicion', '').strip()

        if not estab or not expedicion:
            return Response({'error': 'Se requieren establecimiento y expedicion.'}, status=400)

        hoy = timezone.localtime().date()
        qs = Timbrado.objects.filter(
            punto_sucursal=estab,
            punto_expedicion=expedicion,
            inicio_vigencia__lte=hoy,
            fin_vigencia__gte=hoy,
            is_deleted=False,
        )
        if not qs.exists():
            return Response({'error': 'No hay timbrado vigente para ese punto de emisión.'}, status=404)

        timbrado = qs.order_by('-inicio_vigencia').first()

        max_usado = VentaFactCab.objects.filter(
            timbrado=timbrado, is_deleted=False,
        ).aggregate(max_nro=Max('nro_comprobante'))['max_nro']

        siguiente = (max_usado if max_usado is not None else timbrado.nro_desde - 1) + 1

        if siguiente > timbrado.nro_hasta:
            return Response({'error': 'El timbrado ha alcanzado el límite de comprobantes.'}, status=400)

        return Response({
            'timbrado_id':  timbrado.id,
            'nro_timbrado': timbrado.nro_timbrado,
            'siguiente':    siguiente,
        })

    @action(detail=True, methods=['get'], url_path='pdf', permission_classes=[AllowAny])
    def pdf(self, request, pk=None):
        try:
            factura = (
                VentaFactCab.objects
                .select_related('persona', 'timbrado')
                .prefetch_related('detalle__prs', 'cobranza')
                .get(pk=pk, is_deleted=False)
            )
        except VentaFactCab.DoesNotExist:
            return HttpResponse('Factura no encontrada.', status=404)

        is_anulado = factura.is_anulado

        estab = factura.establecimiento or str(factura.timbrado.punto_sucursal).zfill(3)
        expd  = factura.expedicion or str(factura.timbrado.punto_expedicion).zfill(3)
        nro_formateado = f'{estab}-{expd}-{str(factura.nro_comprobante).zfill(7)}'

        detalle = list(factura.detalle.filter(is_deleted=False).select_related('prs'))

        for item in detalle:
            try:
                from decimal import Decimal as D
                item.precio_unitario = (D(str(item.monto)) / D(str(item.cantidad))).quantize(D('1'))
            except Exception:
                item.precio_unitario = item.monto

        MIN_FILAS    = 8
        filas_vacias = range(max(0, MIN_FILAS - len(detalle)))

        sub_10     = (factura.grav_10 or Decimal('0')) + (factura.iva_10 or Decimal('0'))
        sub_5      = (factura.grav_5  or Decimal('0')) + (factura.iva_5  or Decimal('0'))
        sub_exenta = (factura.monto_total or Decimal('0')) - (factura.total_gravada or Decimal('0')) - (factura.total_iva or Decimal('0'))

        contexto = {
            'factura':        factura,
            'detalle':        detalle,
            'nro_formateado': nro_formateado,
            'condicion':      'Contado' if factura.condicion_vta else 'Crédito',
            'filas_vacias':   filas_vacias,
            'sub_10':         sub_10,
            'sub_5':          sub_5,
            'sub_exenta':     sub_exenta,
            'is_anulado':     is_anulado,
        }

        html = render_to_string('informes/factura_print.html', contexto, request=request)

        try:
            import weasyprint
            pdf_bytes = weasyprint.HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
        except Exception as e:
            return HttpResponse(f'Error generando PDF: {e}', status=500)

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="factura_{nro_formateado}.pdf"'
        return response

    def _qs_con_filtros(self, request):
        from django.db.models import Q
        qs          = VentaFactCab.objects.filter(is_deleted=False).select_related('persona', 'timbrado')
        search      = request.query_params.get('search', '').strip()
        fecha_desde = request.query_params.get('fecha_desde', '').strip()
        fecha_hasta = request.query_params.get('fecha_hasta', '').strip()
        condicion   = request.query_params.get('condicion_vta', '').strip()
        persona_id  = request.query_params.get('persona', '').strip()
        if search:
            qs = qs.filter(
                Q(persona__razon_social__icontains=search) |
                Q(persona__nro_documento__icontains=search) |
                Q(nro_comprobante__icontains=search)
            )
        if persona_id:
            qs = qs.filter(persona_id=persona_id)
        if fecha_desde:
            qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha__lte=fecha_hasta)
        if condicion in ('true', 'false'):
            qs = qs.filter(condicion_vta=condicion == 'true')
        return qs.order_by('persona__razon_social', '-fecha', '-fecha_creacion')

    def _filtros_str(self, request):
        partes      = []
        search      = request.query_params.get('search', '').strip()
        fecha_desde = request.query_params.get('fecha_desde', '').strip()
        fecha_hasta = request.query_params.get('fecha_hasta', '').strip()
        condicion   = request.query_params.get('condicion_vta', '').strip()
        persona_id  = request.query_params.get('persona', '').strip()
        if persona_id:
            try:
                from apps.administracion.persona.models import Persona as PersonaModel
                p = PersonaModel.objects.get(pk=persona_id)
                partes.append(f'Cliente: {p.razon_social}')
            except Exception:
                pass
        if search:
            partes.append(f'Búsqueda: {search}')
        if fecha_desde:
            partes.append(f'Desde: {fecha_desde}')
        if fecha_hasta:
            partes.append(f'Hasta: {fecha_hasta}')
        if condicion == 'true':
            partes.append('Condición: Contado')
        elif condicion == 'false':
            partes.append('Condición: Crédito')
        return ' · '.join(partes) if partes else ''

    def _agrupar_por_cliente(self, qs):
        from collections import OrderedDict
        grupos_dict = OrderedDict()
        for f in qs:
            key = f.persona_id
            estab = f.establecimiento or str(f.timbrado.punto_sucursal).zfill(3)
            expd  = f.expedicion or str(f.timbrado.punto_expedicion).zfill(3)
            if key not in grupos_dict:
                grupos_dict[key] = {
                    'cliente':   f.persona.razon_social,
                    'documento': f.persona.nro_documento,
                    'facturas':  [],
                    'subtotal':  Decimal('0'),
                }
            grupos_dict[key]['facturas'].append({
                'nro':        len(grupos_dict[key]['facturas']) + 1,
                'comprobante': f'{estab}-{expd}-{str(f.nro_comprobante).zfill(7)}',
                'fecha':      f.fecha,
                'condicion':  'Contado' if f.condicion_vta else 'Crédito',
                'monto_total': f.monto_total or Decimal('0'),
            })
            grupos_dict[key]['subtotal'] += f.monto_total or Decimal('0')
        return list(grupos_dict.values())

    def _datos_control(self, request):
        from datetime import date as date_type
        fecha_desde = request.query_params.get('fecha_desde', '').strip()
        fecha_hasta = request.query_params.get('fecha_hasta', '').strip()

        todas = list(
            VentaFactCab.objects
            .filter(is_deleted=False)
            .select_related('persona', 'timbrado')
        )

        timbrado_info = {}
        for f in todas:
            tid = f.timbrado_id
            if tid not in timbrado_info:
                timbrado_info[tid] = {'timbrado': f.timbrado, 'nros': set(), 'max_nro': 0}
            timbrado_info[tid]['nros'].add(f.nro_comprobante)
            if f.nro_comprobante > timbrado_info[tid]['max_nro']:
                timbrado_info[tid]['max_nro'] = f.nro_comprobante

        anuladas = [f for f in todas if f.is_anulado]
        if fecha_desde:
            anuladas = [f for f in anuladas if str(f.fecha) >= fecha_desde]
        if fecha_hasta:
            anuladas = [f for f in anuladas if str(f.fecha) <= fecha_hasta]

        SENTINEL = date_type.max
        filas = []

        for f in anuladas:
            estab = f.establecimiento or str(f.timbrado.punto_sucursal).zfill(3)
            expd  = f.expedicion or str(f.timbrado.punto_expedicion).zfill(3)
            filas.append({
                'tipo':           'anulada',
                'comprobante':    f'{estab}-{expd}-{str(f.nro_comprobante).zfill(7)}',
                'nro_comprobante': f.nro_comprobante,
                'fecha':          f.fecha,
                'cliente':        f.persona.razon_social,
                'condicion':      'Contado' if f.condicion_vta else 'Crédito',
                'timbrado_nro':   f.timbrado.nro_timbrado,
                '_sort':          (f.fecha, f.nro_comprobante),
            })

        for info in timbrado_info.values():
            t       = info['timbrado']
            nros    = info['nros']
            max_nro = info['max_nro']
            estab   = str(t.punto_sucursal).zfill(3)
            expd    = str(t.punto_expedicion).zfill(3)
            for nro in range(t.nro_desde, max_nro + 1):
                if nro not in nros:
                    filas.append({
                        'tipo':           'no_emitida',
                        'comprobante':    f'{estab}-{expd}-{str(nro).zfill(7)}',
                        'nro_comprobante': nro,
                        'fecha':          None,
                        'cliente':        '—',
                        'condicion':      '—',
                        'timbrado_nro':   t.nro_timbrado,
                        '_sort':          (SENTINEL, nro),
                    })

        filas.sort(key=lambda x: x['_sort'])
        for i, f in enumerate(filas, 1):
            f['nro'] = i

        total_anuladas    = sum(1 for f in filas if f['tipo'] == 'anulada')
        total_no_emitidas = sum(1 for f in filas if f['tipo'] == 'no_emitida')
        return filas, fecha_desde, fecha_hasta, total_anuladas, total_no_emitidas

    @action(detail=False, methods=['get'], url_path='reporte-control-pdf')
    def reporte_control_pdf(self, request):
        filas, fecha_desde, fecha_hasta, total_anuladas, total_no_emitidas = self._datos_control(request)
        hoy = timezone.localtime().date()
        partes = []
        if fecha_desde: partes.append(f'Desde: {fecha_desde}')
        if fecha_hasta: partes.append(f'Hasta: {fecha_hasta}')
        filtros_str = ' · '.join(partes)
        contexto = {
            'filas':             filas,
            'total':             len(filas),
            'total_anuladas':    total_anuladas,
            'total_no_emitidas': total_no_emitidas,
            'fecha':             hoy,
            'filtros_str':       filtros_str,
        }
        html = render_to_string('informes/factura_control.html', contexto, request=request)
        try:
            import weasyprint
            pdf_bytes = weasyprint.HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
        except Exception as e:
            return HttpResponse(f'Error generando PDF: {e}', status=500)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="control_comprobantes.pdf"'
        return response

    @action(detail=False, methods=['get'], url_path='reporte-control-excel')
    def reporte_control_excel(self, request):
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        filas, fecha_desde, fecha_hasta, total_anuladas, total_no_emitidas = self._datos_control(request)
        hoy = timezone.localtime().date()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Control Comprobantes'

        fill_header    = PatternFill('solid', fgColor='1A3A5C')
        fill_anulada   = PatternFill('solid', fgColor='FEE2E2')
        fill_no_emit   = PatternFill('solid', fgColor='FEF9C3')
        fill_par       = PatternFill('solid', fgColor='F8FAFC')
        font_header    = Font(color='FFFFFF', bold=True, size=10)
        font_titulo    = Font(color='1A3A5C', bold=True, size=13)
        font_meta      = Font(color='555555', size=9)
        font_anulada   = Font(color='991B1B', size=10)
        font_no_emit   = Font(color='854D0E', size=10)
        thin           = Border(bottom=Side(style='thin', color='E8EDF2'))
        align_c        = Alignment(horizontal='center', vertical='center')
        align_l        = Alignment(horizontal='left',   vertical='center')

        COLS = 7
        fila = 1

        ws.merge_cells(f'A{fila}:{get_column_letter(COLS)}{fila}')
        c = ws.cell(fila, 1, 'Clínica Lichi — Control de Comprobantes')
        c.font = font_titulo; c.alignment = align_l
        ws.row_dimensions[fila].height = 20
        fila += 1

        partes = []
        if fecha_desde: partes.append(f'Desde: {fecha_desde}')
        if fecha_hasta: partes.append(f'Hasta: {fecha_hasta}')
        meta = ((' · '.join(partes) + ' · ') if partes else '') + f"Generado: {hoy.strftime('%d/%m/%Y')} · Anuladas: {total_anuladas} · No emitidas: {total_no_emitidas}"
        ws.merge_cells(f'A{fila}:{get_column_letter(COLS)}{fila}')
        c = ws.cell(fila, 1, meta)
        c.font = font_meta; c.alignment = align_l
        fila += 2

        headers = ['#', 'Tipo', 'Comprobante', 'Fecha', 'Cliente', 'Condición', 'Timbrado']
        for col, txt in enumerate(headers, 1):
            c = ws.cell(fila, col, txt)
            c.fill = fill_header; c.font = font_header
            c.alignment = align_c if col in (1, 3, 4, 6, 7) else align_l
        ws.row_dimensions[fila].height = 18
        fila += 1

        for i, row in enumerate(filas, 1):
            es_anulada  = row['tipo'] == 'anulada'
            es_no_emit  = row['tipo'] == 'no_emitida'
            fill        = fill_anulada if es_anulada else (fill_no_emit if es_no_emit else (fill_par if i % 2 == 0 else None))
            font        = font_anulada if es_anulada else (font_no_emit if es_no_emit else None)
            tipo_label  = 'Anulada' if es_anulada else 'No emitida'
            fecha_val   = row['fecha'].strftime('%d/%m/%Y') if row['fecha'] else '—'
            vals = [row['nro'], tipo_label, row['comprobante'], fecha_val, row['cliente'], row['condicion'], str(row['timbrado_nro'])]
            for col, val in enumerate(vals, 1):
                c = ws.cell(fila, col, val)
                if fill: c.fill = fill
                if font: c.font = font
                c.border    = thin
                c.alignment = align_c if col in (1, 3, 4, 6, 7) else align_l
            ws.row_dimensions[fila].height = 15
            fila += 1

        for i, ancho in enumerate([5, 14, 20, 12, 36, 12, 16], 1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        buf = BytesIO()
        wb.save(buf); buf.seek(0)
        fecha_str = hoy.strftime('%Y%m%d')
        response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="control_comprobantes_{fecha_str}.xlsx"'
        return response

    @action(detail=False, methods=['get'], url_path='dashboard-mensual')
    def dashboard_mensual(self, request):
        from calendar import monthrange
        from django.db.models.functions import ExtractYear, ExtractMonth

        MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        MESES_CORTO = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                       'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

        hoy        = timezone.localtime().date()
        ultimo_dia = monthrange(hoy.year, hoy.month)[1]
        primer_dia = hoy.replace(day=1)
        ultimo     = hoy.replace(day=ultimo_dia)

        qs_mes = list(
            VentaFactCab.objects.filter(
                is_deleted=False,
                fecha__gte=primer_dia,
                fecha__lte=ultimo,
            ).select_related('persona')
        )

        hoy_facts       = [f for f in qs_mes if f.fecha == hoy]
        hoy_activas     = [f for f in hoy_facts if not f.is_anulado]
        hoy_contado     = [f for f in hoy_activas if f.condicion_vta]
        hoy_credito     = [f for f in hoy_activas if not f.condicion_vta]
        stats_hoy = {
            'total':         len(hoy_activas),
            'anuladas':      len([f for f in hoy_facts if f.is_anulado]),
            'contado':       len(hoy_contado),
            'credito':       len(hoy_credito),
            'monto_total':   float(sum((f.monto_total or Decimal('0')) for f in hoy_activas)),
            'monto_contado': float(sum((f.monto_total or Decimal('0')) for f in hoy_contado)),
            'monto_credito': float(sum((f.monto_total or Decimal('0')) for f in hoy_credito)),
        }

        activas_mes   = [f for f in qs_mes if not f.is_anulado]
        total_mes     = len(activas_mes)
        monto_mes     = sum((f.monto_total or Decimal('0')) for f in activas_mes)
        monto_contado = sum((f.monto_total or Decimal('0')) for f in activas_mes if f.condicion_vta)
        monto_credito = sum((f.monto_total or Decimal('0')) for f in activas_mes if not f.condicion_vta)
        totales_mes = {
            'total_facturas':  total_mes,
            'monto_total':     float(monto_mes),
            'anuladas':        len([f for f in qs_mes if f.is_anulado]),
            'contado':         len([f for f in activas_mes if f.condicion_vta]),
            'credito':         len([f for f in activas_mes if not f.condicion_vta]),
            'ticket_promedio': float(monto_mes / total_mes) if total_mes > 0 else 0.0,
            'monto_contado':   float(monto_contado),
            'monto_credito':   float(monto_credito),
        }

        por_dia_dict = {}
        for f in qs_mes:
            k = f.fecha
            if k not in por_dia_dict:
                por_dia_dict[k] = {'total': 0, 'contado': 0, 'credito': 0, 'anuladas': 0, 'monto': Decimal('0')}
            if f.is_anulado:
                por_dia_dict[k]['anuladas'] += 1
            else:
                por_dia_dict[k]['total']   += 1
                por_dia_dict[k]['monto']   += f.monto_total or Decimal('0')
                if f.condicion_vta:
                    por_dia_dict[k]['contado'] += 1
                else:
                    por_dia_dict[k]['credito'] += 1

        por_dia = []
        for dia in range(1, ultimo_dia + 1):
            fecha = hoy.replace(day=dia)
            d     = por_dia_dict.get(fecha, {'total': 0, 'contado': 0, 'credito': 0, 'anuladas': 0, 'monto': Decimal('0')})
            por_dia.append({
                'fecha':      str(fecha),
                'label':      fecha.strftime('%d/%m'),
                'dia':        dia,
                'total':      d['total'],
                'contado':    d['contado'],
                'credito':    d['credito'],
                'anuladas':   d['anuladas'],
                'monto_total': float(d['monto']),
                'es_futuro':  fecha > hoy,
            })

        por_semana = []
        sem = 1
        for inicio in range(0, ultimo_dia, 7):
            dias = list(range(inicio + 1, min(inicio + 8, ultimo_dia + 1)))
            sem_dias = [por_dia[d - 1] for d in dias]
            pasados  = [d for d in sem_dias if not d['es_futuro']]
            por_semana.append({
                'semana':      sem,
                'label':       f'Sem {sem} ({dias[0]}–{dias[-1]})',
                'label_corto': f'S{sem}',
                'total':       sum(d['total']    for d in pasados),
                'contado':     sum(d['contado']  for d in pasados),
                'credito':     sum(d['credito']  for d in pasados),
                'monto_total': sum(d['monto_total'] for d in pasados),
                'es_futuro':   hoy.replace(day=dias[0]) > hoy,
            })
            sem += 1

        anos_meses = []
        for i in range(5, -1, -1):
            m = hoy.month - i
            a = hoy.year
            while m <= 0:
                m += 12; a -= 1
            anos_meses.append((a, m))

        filtro_tend = Q()
        for a, m in anos_meses:
            filtro_tend |= Q(fecha__year=a, fecha__month=m)

        tend_qs = (
            VentaFactCab.objects.filter(is_deleted=False, is_anulado=False)
            .filter(filtro_tend)
            .annotate(anio=ExtractYear('fecha'), mes=ExtractMonth('fecha'))
            .values('anio', 'mes')
            .annotate(total=Count('id'), monto=Sum('monto_total'))
        )
        raw_tend = {(r['anio'], r['mes']): r for r in tend_qs}

        por_mes = [
            {
                'mes':         f'{a}-{m:02d}',
                'label':       f'{MESES[m - 1]} {a}',
                'label_corto': MESES_CORTO[m - 1],
                'total':       raw_tend.get((a, m), {}).get('total', 0),
                'monto_total': float(raw_tend.get((a, m), {}).get('monto', 0) or 0),
            }
            for a, m in anos_meses
        ]

        clientes_dict = {}
        for f in activas_mes:
            pid = f.persona_id
            if pid not in clientes_dict:
                clientes_dict[pid] = {'cliente': f.persona.razon_social, 'total_facturas': 0, 'monto': Decimal('0')}
            clientes_dict[pid]['total_facturas'] += 1
            clientes_dict[pid]['monto'] += f.monto_total or Decimal('0')

        top_clientes = sorted(clientes_dict.values(), key=lambda x: x['monto'], reverse=True)[:5]
        top_clientes = [
            {'cliente': c['cliente'], 'total_facturas': c['total_facturas'], 'monto_total': float(c['monto'])}
            for c in top_clientes
        ]

        return Response({
            'mes_label':    f'{MESES[hoy.month - 1]} {hoy.year}',
            'hoy':          str(hoy),
            'ultimo_dia':   ultimo_dia,
            'stats_hoy':    stats_hoy,
            'totales_mes':  totales_mes,
            'por_dia':      por_dia,
            'por_semana':   por_semana,
            'por_mes':      por_mes,
            'top_clientes': top_clientes,
        })

    def _datos_estado_cuenta(self, request):
        from collections import OrderedDict
        from apps.finanzas.estadocuenta.models import CtaCobrar

        modo               = request.query_params.get('modo', 'detallado').lower()
        usar_rango         = request.query_params.get('usar_rango', 'true').lower() == 'true'
        fecha_desde        = request.query_params.get('fecha_desde', '').strip()
        fecha_hasta        = request.query_params.get('fecha_hasta', '').strip()
        persona_id         = request.query_params.get('persona', '').strip()
        incluir_saldo_cero = request.query_params.get('incluir_saldo_cero', 'false').lower() == 'true'

        qs = (
            CtaCobrar.objects
            .filter(is_deleted=False, vfc__is_deleted=False, vfc__is_anulado=False)
            .select_related('vfc', 'vfc__persona', 'vfc__timbrado')
            .order_by('vfc__persona__razon_social', 'vfc__fecha', 'nro_cuota')
        )

        if usar_rango and fecha_desde:
            qs = qs.filter(vfc__fecha__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(vfc__fecha__lte=fecha_hasta)
        if persona_id:
            qs = qs.filter(vfc__persona_id=persona_id)
        if not incluir_saldo_cero:
            qs = qs.filter(saldo__gt=0)

        grupos = OrderedDict()
        for cuota in qs:
            pid = cuota.vfc.persona_id
            if pid not in grupos:
                grupos[pid] = {
                    'cliente':      cuota.vfc.persona.razon_social,
                    'documento':    cuota.vfc.persona.nro_documento,
                    'cuotas':       [],
                    'total_monto':  Decimal('0'),
                    'total_saldo':  Decimal('0'),
                }
            estab       = cuota.vfc.establecimiento or str(cuota.vfc.timbrado.punto_sucursal).zfill(3)
            expd        = cuota.vfc.expedicion or str(cuota.vfc.timbrado.punto_expedicion).zfill(3)
            comprobante = f'{estab}-{expd}-{str(cuota.vfc.nro_comprobante).zfill(7)}'
            grupos[pid]['cuotas'].append({
                'comprobante':      comprobante,
                'fecha_factura':    cuota.vfc.fecha,
                'fecha_vencimiento': cuota.fecha_vencimiento,
                'nro_cuota':        cuota.nro_cuota,
                'cant_cuota':       cuota.cant_cuota,
                'monto_cuota':      cuota.monto_cuota,
                'saldo':            cuota.saldo,
            })
            grupos[pid]['total_monto'] += cuota.monto_cuota
            grupos[pid]['total_saldo'] += cuota.saldo

        if modo == 'resumido' and not incluir_saldo_cero:
            grupos = OrderedDict((k, v) for k, v in grupos.items() if v['total_saldo'] > 0)

        gran_total_monto = sum(g['total_monto'] for g in grupos.values())
        gran_total_saldo = sum(g['total_saldo'] for g in grupos.values())

        partes = []
        if usar_rango and fecha_desde:
            partes.append(f'Desde: {fecha_desde}')
        if fecha_hasta:
            partes.append(f'Hasta: {fecha_hasta}')
        if not usar_rango and not fecha_hasta:
            partes.append('Sin filtro de fecha')
        if persona_id:
            try:
                from apps.administracion.persona.models import Persona as PM
                p = PM.objects.get(pk=persona_id)
                partes.append(f'Cliente: {p.razon_social}')
            except Exception:
                pass
        if incluir_saldo_cero:
            partes.append('Incluye saldo cero')
        filtros_str = ' · '.join(partes) if partes else 'Sin filtros de fecha'

        return list(grupos.values()), modo, filtros_str, gran_total_monto, gran_total_saldo

    @action(detail=False, methods=['get'], url_path='estado-cuenta-pdf')
    def estado_cuenta_pdf(self, request):
        grupos, modo, filtros_str, gran_total_monto, gran_total_saldo = self._datos_estado_cuenta(request)
        hoy = timezone.localtime().date()

        contexto = {
            'grupos':            grupos,
            'modo':              modo,
            'filtros_str':       filtros_str,
            'gran_total_monto':  gran_total_monto,
            'gran_total_saldo':  gran_total_saldo,
            'total_clientes':    len(grupos),
            'fecha':             hoy,
        }
        html = render_to_string('informes/estado_cuenta.html', contexto, request=request)
        try:
            import weasyprint
            pdf_bytes = weasyprint.HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
        except Exception as e:
            return HttpResponse(f'Error generando PDF: {e}', status=500)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="estado_cuenta_{hoy.strftime("%Y%m%d")}.pdf"'
        return response

    @action(detail=False, methods=['get'], url_path='estado-cuenta-excel')
    def estado_cuenta_excel(self, request):
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        grupos, modo, filtros_str, gran_total_monto, gran_total_saldo = self._datos_estado_cuenta(request)
        hoy = timezone.localtime().date()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Estado de Cuenta'

        fill_hdr    = PatternFill('solid', fgColor='1A3A5C')
        fill_grupo  = PatternFill('solid', fgColor='EEF2F7')
        fill_sub    = PatternFill('solid', fgColor='DBEAFE')
        fill_total  = PatternFill('solid', fgColor='1A3A5C')
        fill_par    = PatternFill('solid', fgColor='F8FAFC')
        font_hdr    = Font(color='FFFFFF', bold=True, size=10)
        font_titulo = Font(color='1A3A5C', bold=True, size=13)
        font_meta   = Font(color='555555', size=9)
        font_grupo  = Font(color='1A3A5C', bold=True, size=10)
        font_sub    = Font(color='1E3A5F', bold=True, size=9)
        font_total  = Font(color='FFFFFF', bold=True, size=10)
        thin        = Border(bottom=Side(style='thin', color='E8EDF2'))
        align_c     = Alignment(horizontal='center', vertical='center')
        align_r     = Alignment(horizontal='right',  vertical='center')
        align_l     = Alignment(horizontal='left',   vertical='center')

        fila = 1
        modo_label = 'Detallado' if modo == 'detallado' else 'Resumido'

        if modo == 'detallado':
            COLS = 7
            HEADERS = ['Comprobante', 'F. Factura', 'F. Vencimiento', 'Cuota', 'Monto Cuota', 'Saldo', 'Estado']
            col_w   = [18, 12, 14, 8, 16, 16, 12]
        else:
            COLS = 5
            HEADERS = ['Cliente', 'Documento', 'Cant. Cuotas', 'Total Facturado', 'Saldo Pendiente']
            col_w   = [35, 16, 13, 18, 18]

        ws.merge_cells(f'A{fila}:{get_column_letter(COLS)}{fila}')
        c = ws.cell(fila, 1, f'Clínica Lichi — Estado de Cuenta ({modo_label})')
        c.font = font_titulo; c.alignment = align_l
        ws.row_dimensions[fila].height = 20; fila += 1

        ws.merge_cells(f'A{fila}:{get_column_letter(COLS)}{fila}')
        meta = f'Generado: {hoy.strftime("%d/%m/%Y")} · {filtros_str} · Clientes: {len(grupos)}'
        c = ws.cell(fila, 1, meta); c.font = font_meta; c.alignment = align_l
        ws.row_dimensions[fila].height = 14; fila += 1; fila += 1

        if modo == 'detallado':
            for i, h in enumerate(HEADERS, 1):
                c = ws.cell(fila, i, h)
                c.font = font_hdr; c.fill = fill_hdr
                c.alignment = align_c if i > 1 else align_l
            ws.row_dimensions[fila].height = 16; fila += 1

            for grupo in grupos:
                ws.merge_cells(f'A{fila}:{get_column_letter(COLS)}{fila}')
                c = ws.cell(fila, 1, f'{grupo["cliente"]}  —  {grupo["documento"]}')
                c.font = font_grupo; c.fill = fill_grupo; c.alignment = align_l
                ws.row_dimensions[fila].height = 14; fila += 1

                for j, cuota in enumerate(grupo['cuotas']):
                    fill = fill_par if j % 2 == 0 else None
                    vals = [
                        cuota['comprobante'],
                        cuota['fecha_factura'].strftime('%d/%m/%Y'),
                        cuota['fecha_vencimiento'].strftime('%d/%m/%Y'),
                        f'{cuota["nro_cuota"]}/{cuota["cant_cuota"]}',
                        int(cuota['monto_cuota']),
                        int(cuota['saldo']),
                        'Pagado' if cuota['saldo'] == 0 else 'Pendiente',
                    ]
                    aligns = [align_c, align_c, align_c, align_c, align_r, align_r, align_c]
                    for ci, (v, al) in enumerate(zip(vals, aligns), 1):
                        cell = ws.cell(fila, ci, v)
                        cell.alignment = al; cell.border = thin
                        if fill: cell.fill = fill
                        if ci in (5, 6): cell.number_format = '#,##0'
                    ws.row_dimensions[fila].height = 13; fila += 1

                ws.merge_cells(f'A{fila}:{get_column_letter(COLS - 2)}{fila}')
                ws.cell(fila, 1, 'Subtotal cliente').font = font_sub
                ws.cell(fila, 1).fill = fill_sub; ws.cell(fila, 1).alignment = align_r
                c5 = ws.cell(fila, COLS - 1, int(grupo['total_monto']))
                c5.font = font_sub; c5.fill = fill_sub; c5.alignment = align_r; c5.number_format = '#,##0'
                c6 = ws.cell(fila, COLS, int(grupo['total_saldo']))
                c6.font = font_sub; c6.fill = fill_sub; c6.alignment = align_r; c6.number_format = '#,##0'
                ws.row_dimensions[fila].height = 13; fila += 1
        else:
            for i, h in enumerate(HEADERS, 1):
                c = ws.cell(fila, i, h)
                c.font = font_hdr; c.fill = fill_hdr
                c.alignment = align_r if i > 2 else align_l
            ws.row_dimensions[fila].height = 16; fila += 1

            for j, grupo in enumerate(grupos):
                fill = fill_par if j % 2 == 0 else None
                vals = [grupo['cliente'], grupo['documento'], len(grupo['cuotas']),
                        int(grupo['total_monto']), int(grupo['total_saldo'])]
                aligns = [align_l, align_c, align_c, align_r, align_r]
                for ci, (v, al) in enumerate(zip(vals, aligns), 1):
                    cell = ws.cell(fila, ci, v)
                    cell.alignment = al; cell.border = thin
                    if fill: cell.fill = fill
                    if ci in (4, 5): cell.number_format = '#,##0'
                ws.row_dimensions[fila].height = 13; fila += 1

        fila += 1
        ws.merge_cells(f'A{fila}:{get_column_letter(COLS - 2)}{fila}')
        c = ws.cell(fila, 1, 'TOTAL GENERAL')
        c.font = font_total; c.fill = fill_total; c.alignment = align_l
        cm = ws.cell(fila, COLS - 1, int(gran_total_monto))
        cm.font = font_total; cm.fill = fill_total; cm.alignment = align_r; cm.number_format = '#,##0'
        cs = ws.cell(fila, COLS, int(gran_total_saldo))
        cs.font = font_total; cs.fill = fill_total; cs.alignment = align_r; cs.number_format = '#,##0'
        ws.row_dimensions[fila].height = 16

        for ci, w in enumerate(col_w, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        buf = BytesIO()
        wb.save(buf); buf.seek(0)
        response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="estado_cuenta_{hoy.strftime("%Y%m%d")}.xlsx"'
        return response

    def _datos_extracto(self, request):
        from collections import OrderedDict
        from apps.finanzas.cobranzas.models import Cobranza, CobranzaDet
        from apps.finanzas.estadocuenta.models import CtaCobrar
        from django.db.models import Sum

        usar_rango          = request.query_params.get('usar_rango', 'true').lower() == 'true'
        fecha_desde         = request.query_params.get('fecha_desde', '').strip()
        fecha_hasta         = request.query_params.get('fecha_hasta', '').strip()
        persona_id          = request.query_params.get('persona', '').strip()
        incluir_saldo_cero  = request.query_params.get('incluir_saldo_cero', 'false').lower() == 'true'
        agrupar_por_factura = request.query_params.get('agrupar_por_factura', 'false').lower() == 'true'

        qs_fact = (
            VentaFactCab.objects
            .filter(is_deleted=False, is_anulado=False, condicion_vta=False, persona__is_deleted=False)
            .select_related('persona', 'timbrado')
        )
        if usar_rango and fecha_desde:
            qs_fact = qs_fact.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            qs_fact = qs_fact.filter(fecha__lte=fecha_hasta)
        if persona_id:
            qs_fact = qs_fact.filter(persona_id=persona_id)

        grupos = OrderedDict()

        if agrupar_por_factura:
            fact_list = list(qs_fact.order_by('persona__razon_social', 'persona_id', 'fecha', 'nro_comprobante'))
            fact_ids  = [f.id for f in fact_list]

            det_qs = (
                CobranzaDet.objects
                .filter(is_deleted=False, cta_cobrar__vfc_id__in=fact_ids, cta_cobrar__is_deleted=False,
                        cobranza__is_deleted=False)
                .select_related('cobranza', 'cta_cobrar')
                .order_by('cta_cobrar__vfc_id', 'cobranza__fecha', 'cobranza_id')
            )
            cobros_por_fact = {}
            for det in det_qs:
                vfc_id = det.cta_cobrar.vfc_id
                cob_id = det.cobranza_id
                cobros_por_fact.setdefault(vfc_id, OrderedDict())
                if cob_id not in cobros_por_fact[vfc_id]:
                    cobros_por_fact[vfc_id][cob_id] = {
                        'fecha':        det.cobranza.fecha,
                        'comprobante':  str(det.cobranza.comprobante_nro or '').zfill(7),
                        'monto':        Decimal('0'),
                    }
                cobros_por_fact[vfc_id][cob_id]['monto'] += det.monto_pagado

            saldo_qs = (
                CtaCobrar.objects
                .filter(is_deleted=False, vfc_id__in=fact_ids)
                .values('vfc_id')
                .annotate(saldo_total=Sum('saldo'))
            )
            saldo_por_fact = {r['vfc_id']: r['saldo_total'] or Decimal('0') for r in saldo_qs}

            for f in fact_list:
                pid = f.persona_id
                if pid not in grupos:
                    grupos[pid] = {
                        'cliente':        f.persona.razon_social,
                        'documento':      f.persona.nro_documento,
                        'facturas':       [],
                        'total_debito':   Decimal('0'),
                        'total_credito':  Decimal('0'),
                        'saldo_final':    Decimal('0'),
                    }
                estab  = f.establecimiento or str(f.timbrado.punto_sucursal).zfill(3)
                expd   = f.expedicion or str(f.timbrado.punto_expedicion).zfill(3)
                comprobante   = f'{estab}-{expd}-{str(f.nro_comprobante).zfill(7)}'
                cobros_list   = list(cobros_por_fact.get(f.id, {}).values())
                total_cobrado = sum(c['monto'] for c in cobros_list)
                saldo         = saldo_por_fact.get(f.id, f.monto_total)
                grupos[pid]['facturas'].append({
                    'comprobante':    comprobante,
                    'fecha':          f.fecha,
                    'monto_total':    f.monto_total,
                    'cobros':         cobros_list,
                    'total_cobrado':  total_cobrado,
                    'saldo':          saldo,
                })
                grupos[pid]['total_debito']  += f.monto_total
                grupos[pid]['total_credito'] += total_cobrado
                grupos[pid]['saldo_final']   += saldo
        else:
            qs_cob = (
                Cobranza.objects
                .filter(is_deleted=False, persona__is_deleted=False)
                .select_related('persona')
            )
            if usar_rango and fecha_desde:
                qs_cob = qs_cob.filter(fecha__gte=fecha_desde)
            if fecha_hasta:
                qs_cob = qs_cob.filter(fecha__lte=fecha_hasta)
            if persona_id:
                qs_cob = qs_cob.filter(persona_id=persona_id)

            for f in qs_fact:
                pid = f.persona_id
                if pid not in grupos:
                    grupos[pid] = {
                        'cliente':       f.persona.razon_social,
                        'documento':     f.persona.nro_documento,
                        'movimientos':   [],
                        'total_debito':  Decimal('0'),
                        'total_credito': Decimal('0'),
                        'saldo_final':   Decimal('0'),
                    }
                estab  = f.establecimiento or str(f.timbrado.punto_sucursal).zfill(3)
                expd   = f.expedicion or str(f.timbrado.punto_expedicion).zfill(3)
                grupos[pid]['movimientos'].append({
                    'tipo':        'factura',
                    'fecha':       f.fecha,
                    'comprobante': f'{estab}-{expd}-{str(f.nro_comprobante).zfill(7)}',
                    'debito':      f.monto_total,
                    'credito':     Decimal('0'),
                })
                grupos[pid]['total_debito'] += f.monto_total

            for c in qs_cob:
                pid = c.persona_id
                if pid not in grupos:
                    grupos[pid] = {
                        'cliente':       c.persona.razon_social,
                        'documento':     c.persona.nro_documento,
                        'movimientos':   [],
                        'total_debito':  Decimal('0'),
                        'total_credito': Decimal('0'),
                        'saldo_final':   Decimal('0'),
                    }
                grupos[pid]['movimientos'].append({
                    'tipo':        'cobranza',
                    'fecha':       c.fecha,
                    'comprobante': str(c.comprobante_nro or '').zfill(7),
                    'debito':      Decimal('0'),
                    'credito':     c.monto,
                })
                grupos[pid]['total_credito'] += c.monto

            for g in grupos.values():
                g['movimientos'].sort(key=lambda m: (m['fecha'], m['tipo']))
                saldo = Decimal('0')
                for m in g['movimientos']:
                    saldo += m['debito'] - m['credito']
                    m['saldo'] = saldo
                g['saldo_final'] = saldo

        if not incluir_saldo_cero:
            grupos = OrderedDict((k, v) for k, v in grupos.items() if v['saldo_final'] > 0)

        grupos = OrderedDict(sorted(grupos.items(), key=lambda x: x[1]['cliente']))

        gran_total_debito  = sum(g['total_debito']  for g in grupos.values())
        gran_total_credito = sum(g['total_credito'] for g in grupos.values())
        gran_total_saldo   = sum(g['saldo_final']   for g in grupos.values())

        partes = []
        if usar_rango and fecha_desde:
            partes.append(f'Desde: {fecha_desde}')
        if fecha_hasta:
            partes.append(f'Hasta: {fecha_hasta}')
        if not usar_rango and not fecha_hasta:
            partes.append('Sin filtro de fecha')
        if persona_id:
            try:
                from apps.administracion.persona.models import Persona as PM
                p = PM.objects.get(pk=persona_id)
                partes.append(f'Cliente: {p.razon_social}')
            except Exception:
                pass
        if incluir_saldo_cero:
            partes.append('Incluye saldo cero')
        filtros_str = ' · '.join(partes) if partes else 'Sin filtros de fecha'

        return (
            list(grupos.values()),
            agrupar_por_factura,
            filtros_str,
            gran_total_debito,
            gran_total_credito,
            gran_total_saldo,
        )

    @action(detail=False, methods=['get'], url_path='extracto-cuenta-pdf')
    def extracto_cuenta_pdf(self, request):
        grupos, agrupar, filtros_str, gran_deb, gran_cred, gran_saldo = self._datos_extracto(request)
        hoy = timezone.localtime().date()
        contexto = {
            'grupos':           grupos,
            'agrupar':          agrupar,
            'filtros_str':      filtros_str,
            'gran_total_debito': gran_deb,
            'gran_total_credito': gran_cred,
            'gran_total_saldo': gran_saldo,
            'total_clientes':   len(grupos),
            'fecha':            hoy,
        }
        html = render_to_string('informes/extracto_cuenta.html', contexto, request=request)
        try:
            import weasyprint
            pdf_bytes = weasyprint.HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
        except Exception as e:
            return HttpResponse(f'Error generando PDF: {e}', status=500)
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = f'inline; filename="extracto_cuenta_{hoy.strftime("%Y%m%d")}.pdf"'
        return resp

    @action(detail=False, methods=['get'], url_path='extracto-cuenta-excel')
    def extracto_cuenta_excel(self, request):
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        grupos, agrupar, filtros_str, gran_deb, gran_cred, gran_saldo = self._datos_extracto(request)
        hoy = timezone.localtime().date()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Extracto de Cuenta'

        fill_hdr    = PatternFill('solid', fgColor='1A3A5C')
        fill_cli    = PatternFill('solid', fgColor='EEF2F7')
        fill_fact   = PatternFill('solid', fgColor='DBEAFE')
        fill_sub    = PatternFill('solid', fgColor='D1FAE5')
        fill_total  = PatternFill('solid', fgColor='1A3A5C')
        fill_par    = PatternFill('solid', fgColor='F8FAFC')
        font_hdr    = Font(color='FFFFFF', bold=True, size=10)
        font_titulo = Font(color='1A3A5C', bold=True, size=13)
        font_meta   = Font(color='555555', size=9)
        font_cli    = Font(color='1A3A5C', bold=True, size=10)
        font_fact   = Font(color='1E3A5F', bold=True, size=9)
        font_sub    = Font(color='065F46', bold=True, size=9)
        font_total  = Font(color='FFFFFF', bold=True, size=10)
        thin        = Border(bottom=Side(style='thin', color='E8EDF2'))
        align_l     = Alignment(horizontal='left',   vertical='center')
        align_c     = Alignment(horizontal='center', vertical='center')
        align_r     = Alignment(horizontal='right',  vertical='center')

        fila = 1
        ws.merge_cells(f'A{fila}:G{fila}')
        c = ws.cell(fila, 1, 'Clínica Lichi — Extracto de Cuenta Clientes')
        c.font = font_titulo; c.alignment = align_l
        ws.row_dimensions[fila].height = 20; fila += 1

        ws.merge_cells(f'A{fila}:G{fila}')
        modo_str = 'Agrupado por Factura' if agrupar else 'Cronológico por Cliente'
        c = ws.cell(fila, 1, f'Modo: {modo_str}   |   {filtros_str}   |   Generado: {hoy.strftime("%d/%m/%Y")}')
        c.font = font_meta; c.alignment = align_l
        fila += 2

        if agrupar:
            COLS = 5
            HEADERS = ['Comprobante', 'Fecha', 'Tipo', 'Débito (Gs.)', 'Crédito (Gs.)', 'Saldo (Gs.)']
            HEADERS = ['Comprobante', 'Fecha', 'Monto Factura', 'Total Cobrado', 'Saldo']
            col_w   = [20, 13, 17, 17, 14]

            for grupo in grupos:
                ws.merge_cells(f'A{fila}:E{fila}')
                c = ws.cell(fila, 1, f'{grupo["cliente"]}  —  {grupo["documento"]}')
                c.font = font_cli; c.fill = fill_cli; c.alignment = align_l
                ws.row_dimensions[fila].height = 14; fila += 1

                for i, h in enumerate(HEADERS, 1):
                    c = ws.cell(fila, i, h)
                    c.font = font_hdr; c.fill = fill_hdr
                    c.alignment = align_r if i > 2 else align_l
                ws.row_dimensions[fila].height = 14; fila += 1

                for j, fact in enumerate(grupo['facturas']):
                    fill_f = fill_par if j % 2 == 0 else None
                    vals   = [fact['comprobante'], fact['fecha'].strftime('%d/%m/%Y'),
                              int(fact['monto_total']), int(fact['total_cobrado']), int(fact['saldo'])]
                    aligns = [align_l, align_c, align_r, align_r, align_r]
                    for ci, (v, al) in enumerate(zip(vals, aligns), 1):
                        cell = ws.cell(fila, ci, v)
                        cell.fill = fill_fact; cell.alignment = al; cell.border = thin
                        if ci >= 3: cell.number_format = '#,##0'
                    ws.row_dimensions[fila].height = 13; fila += 1

                    for cob in fact['cobros']:
                        c1 = ws.cell(fila, 1, f'  Recibo #{cob["comprobante"]}')
                        c1.alignment = align_l; c1.border = thin
                        c2 = ws.cell(fila, 2, cob['fecha'].strftime('%d/%m/%Y'))
                        c2.alignment = align_c; c2.border = thin
                        ws.cell(fila, 3, '').border = thin
                        c4 = ws.cell(fila, 4, int(cob['monto']))
                        c4.alignment = align_r; c4.border = thin; c4.number_format = '#,##0'
                        ws.cell(fila, 5, '').border = thin
                        ws.row_dimensions[fila].height = 13; fila += 1

                ws.merge_cells(f'A{fila}:B{fila}')
                ws.cell(fila, 1, 'Subtotal cliente').font = font_sub
                ws.cell(fila, 1).fill = fill_sub; ws.cell(fila, 1).alignment = align_r
                c3 = ws.cell(fila, 3, int(grupo['total_debito']))
                c3.font = font_sub; c3.fill = fill_sub; c3.alignment = align_r; c3.number_format = '#,##0'
                c4 = ws.cell(fila, 4, int(grupo['total_credito']))
                c4.font = font_sub; c4.fill = fill_sub; c4.alignment = align_r; c4.number_format = '#,##0'
                c5 = ws.cell(fila, 5, int(grupo['saldo_final']))
                c5.font = font_sub; c5.fill = fill_sub; c5.alignment = align_r; c5.number_format = '#,##0'
                ws.row_dimensions[fila].height = 13; fila += 2

        else:
            HEADERS = ['Fecha', 'Tipo', 'Comprobante', 'Débito (Gs.)', 'Crédito (Gs.)', 'Saldo (Gs.)']
            COLS    = len(HEADERS)
            col_w   = [13, 12, 22, 17, 17, 17]

            for grupo in grupos:
                ws.merge_cells(f'A{fila}:{get_column_letter(COLS)}{fila}')
                c = ws.cell(fila, 1, f'{grupo["cliente"]}  —  {grupo["documento"]}')
                c.font = font_cli; c.fill = fill_cli; c.alignment = align_l
                ws.row_dimensions[fila].height = 14; fila += 1

                for i, h in enumerate(HEADERS, 1):
                    c = ws.cell(fila, i, h)
                    c.font = font_hdr; c.fill = fill_hdr
                    c.alignment = align_r if i > 3 else align_l
                ws.row_dimensions[fila].height = 14; fila += 1

                for j, mov in enumerate(grupo['movimientos']):
                    fill_m = fill_par if j % 2 == 0 else None
                    tipo_label = 'Factura' if mov['tipo'] == 'factura' else 'Cobranza'
                    vals   = [mov['fecha'].strftime('%d/%m/%Y'), tipo_label, mov['comprobante'],
                              int(mov['debito']) if mov['debito'] else '',
                              int(mov['credito']) if mov['credito'] else '',
                              int(mov['saldo'])]
                    aligns = [align_c, align_l, align_l, align_r, align_r, align_r]
                    for ci, (v, al) in enumerate(zip(vals, aligns), 1):
                        cell = ws.cell(fila, ci, v)
                        cell.border = thin; cell.alignment = al
                        if fill_m: cell.fill = fill_m
                        if ci >= 4 and v != '': cell.number_format = '#,##0'
                    ws.row_dimensions[fila].height = 13; fila += 1

                ws.merge_cells(f'A{fila}:{get_column_letter(COLS - 3)}{fila}')
                ws.cell(fila, 1, 'Subtotal cliente').font = font_sub
                ws.cell(fila, 1).fill = fill_sub; ws.cell(fila, 1).alignment = align_r
                c4 = ws.cell(fila, COLS - 2, int(grupo['total_debito']))
                c4.font = font_sub; c4.fill = fill_sub; c4.alignment = align_r; c4.number_format = '#,##0'
                c5 = ws.cell(fila, COLS - 1, int(grupo['total_credito']))
                c5.font = font_sub; c5.fill = fill_sub; c5.alignment = align_r; c5.number_format = '#,##0'
                c6 = ws.cell(fila, COLS, int(grupo['saldo_final']))
                c6.font = font_sub; c6.fill = fill_sub; c6.alignment = align_r; c6.number_format = '#,##0'
                ws.row_dimensions[fila].height = 13; fila += 2

        fila += 1
        COLS_TOT = 5 if agrupar else 6
        ws.merge_cells(f'A{fila}:{get_column_letter(COLS_TOT - 2)}{fila}')
        c = ws.cell(fila, 1, 'GRAN TOTAL')
        c.font = font_total; c.fill = fill_total; c.alignment = align_l
        c_deb = ws.cell(fila, COLS_TOT - 2, int(gran_deb))
        c_deb.font = font_total; c_deb.fill = fill_total; c_deb.alignment = align_r; c_deb.number_format = '#,##0'
        c_cre = ws.cell(fila, COLS_TOT - 1, int(gran_cred))
        c_cre.font = font_total; c_cre.fill = fill_total; c_cre.alignment = align_r; c_cre.number_format = '#,##0'
        c_sal = ws.cell(fila, COLS_TOT, int(gran_saldo))
        c_sal.font = font_total; c_sal.fill = fill_total; c_sal.alignment = align_r; c_sal.number_format = '#,##0'
        ws.row_dimensions[fila].height = 16

        col_w_fin = col_w if agrupar else col_w
        for ci, w in enumerate(col_w_fin, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        buf = BytesIO()
        wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="extracto_cuenta_{hoy.strftime("%Y%m%d")}.xlsx"'
        return resp

    @action(detail=False, methods=['get'], url_path='reporte-pdf')
    def reporte_pdf(self, request):
        qs       = self._qs_con_filtros(request)
        agrupar  = request.query_params.get('agrupar_cliente', '').strip() == 'true'
        hoy      = timezone.localtime().date()
        filtros  = self._filtros_str(request)

        if agrupar:
            grupos        = self._agrupar_por_cliente(qs)
            total         = sum(len(g['facturas']) for g in grupos)
            total_monto   = sum(g['subtotal'] for g in grupos)
            contexto = {
                'agrupado': True, 'grupos': grupos,
                'total': total, 'total_monto': total_monto,
                'fecha': hoy, 'filtros_str': filtros,
            }
        else:
            filas = []
            for i, f in enumerate(qs, 1):
                estab = f.establecimiento or str(f.timbrado.punto_sucursal).zfill(3)
                expd  = f.expedicion or str(f.timbrado.punto_expedicion).zfill(3)
                filas.append({
                    'nro':         i,
                    'comprobante': f'{estab}-{expd}-{str(f.nro_comprobante).zfill(7)}',
                    'fecha':       f.fecha,
                    'cliente':     f.persona.razon_social,
                    'documento':   f.persona.nro_documento,
                    'condicion':   'Contado' if f.condicion_vta else 'Crédito',
                    'monto_total': f.monto_total or Decimal('0'),
                })
            total_monto = sum(r['monto_total'] for r in filas)
            contexto = {
                'agrupado': False, 'filas': filas,
                'total': len(filas), 'total_monto': total_monto,
                'fecha': hoy, 'filtros_str': filtros,
            }

        html = render_to_string('informes/factura_lista.html', contexto, request=request)
        try:
            import weasyprint
            pdf_bytes = weasyprint.HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
        except Exception as e:
            return HttpResponse(f'Error generando PDF: {e}', status=500)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="facturas.pdf"'
        return response

    @action(detail=False, methods=['get'], url_path='reporte-excel')
    def reporte_excel(self, request):
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        qs      = self._qs_con_filtros(request)
        agrupar = request.query_params.get('agrupar_cliente', '').strip() == 'true'
        hoy     = timezone.localtime().date()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Facturas'

        fill_header   = PatternFill('solid', fgColor='1A3A5C')
        fill_grupo    = PatternFill('solid', fgColor='EEF2F7')
        fill_subtotal = PatternFill('solid', fgColor='DBEAFE')
        fill_par      = PatternFill('solid', fgColor='F8FAFC')
        font_header   = Font(color='FFFFFF', bold=True, size=10)
        font_grupo    = Font(color='1A3A5C', bold=True, size=10)
        font_sub      = Font(color='1E3A5F', bold=True, size=9)
        font_titulo   = Font(color='1A3A5C', bold=True, size=13)
        font_meta     = Font(color='555555', size=9)
        thin          = Border(bottom=Side(style='thin', color='E8EDF2'))
        align_c       = Alignment(horizontal='center', vertical='center')
        align_l       = Alignment(horizontal='left',   vertical='center')
        align_r       = Alignment(horizontal='right',  vertical='center')

        COLS = 6
        fila = 1

        ws.merge_cells(f'A{fila}:{get_column_letter(COLS)}{fila}')
        c = ws.cell(fila, 1, 'Clínica Lichi — Listado de Facturas')
        c.font = font_titulo; c.alignment = align_l
        ws.row_dimensions[fila].height = 20
        fila += 1

        ws.merge_cells(f'A{fila}:{get_column_letter(COLS)}{fila}')
        c = ws.cell(fila, 1, f"Generado el {hoy.strftime('%d/%m/%Y')}")
        c.font = font_meta; c.alignment = align_l
        fila += 2

        headers = ['#', 'Nro. Comprobante', 'Fecha', 'Condición', 'Cliente', 'Total (Gs.)']
        if not agrupar:
            headers = ['#', 'Nro. Comprobante', 'Fecha', 'Cliente', 'Documento', 'Condición', 'Total (Gs.)']
            COLS = 7

        for col, txt in enumerate(headers, 1):
            c = ws.cell(fila, col, txt)
            c.fill = fill_header; c.font = font_header
            c.alignment = align_c if col in (1, 3, 4) else align_l
        ws.row_dimensions[fila].height = 18
        fila += 1

        total_monto = Decimal('0')

        if agrupar:
            grupos = self._agrupar_por_cliente(qs)
            for grupo in grupos:
                ws.merge_cells(f'A{fila}:{get_column_letter(COLS)}{fila}')
                c = ws.cell(fila, 1, f"{grupo['cliente']}  ({grupo['documento']})")
                c.fill = fill_grupo; c.font = font_grupo; c.alignment = align_l
                ws.row_dimensions[fila].height = 16
                fila += 1

                for nro, fact in enumerate(grupo['facturas'], 1):
                    es_par = (nro % 2 == 0)
                    vals   = [nro, fact['comprobante'], fact['fecha'], fact['condicion'], grupo['cliente'], int(fact['monto_total'])]
                    for col, val in enumerate(vals, 1):
                        c = ws.cell(fila, col, val)
                        if es_par: c.fill = fill_par
                        c.border    = thin
                        c.alignment = align_c if col in (1, 3, 4) else (align_r if col == 6 else align_l)
                    ws.row_dimensions[fila].height = 15
                    fila += 1

                ws.merge_cells(f'A{fila}:{get_column_letter(COLS - 1)}{fila}')
                c = ws.cell(fila, 1, f"Subtotal {grupo['cliente']}")
                c.fill = fill_subtotal; c.font = font_sub; c.alignment = align_r
                c = ws.cell(fila, COLS, int(grupo['subtotal']))
                c.fill = fill_subtotal; c.font = font_sub; c.alignment = align_r
                ws.row_dimensions[fila].height = 15
                fila += 1
                total_monto += grupo['subtotal']

            for i, ancho in enumerate([5, 18, 12, 12, 32, 16], 1):
                ws.column_dimensions[get_column_letter(i)].width = ancho
        else:
            for i, f in enumerate(qs, 1):
                estab = f.establecimiento or str(f.timbrado.punto_sucursal).zfill(3)
                expd  = f.expedicion or str(f.timbrado.punto_expedicion).zfill(3)
                monto = f.monto_total or Decimal('0')
                total_monto += monto
                vals  = [
                    i, f'{estab}-{expd}-{str(f.nro_comprobante).zfill(7)}',
                    f.fecha, f.persona.razon_social, f.persona.nro_documento,
                    'Contado' if f.condicion_vta else 'Crédito', int(monto),
                ]
                for col, val in enumerate(vals, 1):
                    c = ws.cell(fila, col, val)
                    if i % 2 == 0: c.fill = fill_par
                    c.border    = thin
                    c.alignment = align_c if col in (1, 3, 6) else (align_r if col == 7 else align_l)
                ws.row_dimensions[fila].height = 15
                fila += 1

            for i, ancho in enumerate([5, 18, 12, 32, 16, 12, 16], 1):
                ws.column_dimensions[get_column_letter(i)].width = ancho

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        fecha_str = hoy.strftime('%Y%m%d')
        response  = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="facturas_{fecha_str}.xlsx"'
        return response
