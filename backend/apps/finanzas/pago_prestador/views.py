from decimal import Decimal

from django.db import transaction
from django.db.models import Max, Q
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone

from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.administracion.auditoria.mixins import AuditoriaMixin, _serializar
from apps.core.permissions import IsAdminRole
from apps.forma_pago.models import FormaPago
from apps.finanzas.caja_banco.models import CuentaMcb, MovimientoCajaBanco
from apps.clinica.agenda.models import Agenda
from apps.administracion.persona_rrhh.models import PersonaRRHH

from .models import PagoPrestador, PagoPrestadorDetCobranza
from .serializers import (
    PagoPrestadorListSerializer,
    PagoPrestadorDetalleSerializer,
    PagoPrestadorCreateSerializer,
)


def _siguiente_nro():
    max_nro = PagoPrestador.objects.filter(
        is_deleted=False, nro_comprobante__isnull=False,
    ).aggregate(max_nro=Max('nro_comprobante'))['max_nro']
    return (max_nro or 0) + 1


class PagoPrestadorViewSet(AuditoriaMixin, viewsets.ModelViewSet):
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields   = ['persona_rrhh__persona__razon_social', 'persona_rrhh__persona__nro_documento']
    ordering_fields = ['fecha_pago', 'monto_total', 'fecha_creacion']
    ordering        = ['-fecha_pago', '-fecha_creacion']

    def get_permissions(self):
        if self.action in ('list', 'retrieve', 'siguiente_numero', 'bloques_pendientes', 'medicos_con_pendientes', 'validar_numero', 'recibo_pdf', 'reporte_pdf', 'reporte_excel'):
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsAdminRole()]

    @action(detail=False, methods=['get'], url_path='eliminados')
    def eliminados(self, request):
        qs = PagoPrestador.objects.filter(is_deleted=True).select_related('persona_rrhh__persona')
        return Response(PagoPrestadorListSerializer(qs, many=True).data)

    def get_queryset(self):
        qs = PagoPrestador.objects.filter(is_deleted=False).select_related('persona_rrhh__persona')
        if rrhh_id := self.request.query_params.get('persona_rrhh'):
            qs = qs.filter(persona_rrhh_id=rrhh_id)
        if estado := self.request.query_params.get('estado'):
            qs = qs.filter(estado=estado)
        if desde := self.request.query_params.get('fecha_desde'):
            qs = qs.filter(fecha_pago__gte=desde)
        if hasta := self.request.query_params.get('fecha_hasta'):
            qs = qs.filter(fecha_pago__lte=hasta)
        return qs

    def get_serializer_class(self):
        if self.action == 'create':
            return PagoPrestadorCreateSerializer
        if self.action == 'retrieve':
            return PagoPrestadorDetalleSerializer
        return PagoPrestadorListSerializer

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        ser = PagoPrestadorCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        try:
            medico = PersonaRRHH.objects.get(pk=data['persona_rrhh_id'], is_deleted=False)
        except PersonaRRHH.DoesNotExist:
            raise ValidationError({'persona_rrhh_id': 'Prestador no encontrado.'})

        nro_comprobante = data.get('nro_comprobante') or _siguiente_nro()
        if PagoPrestador.objects.filter(nro_comprobante=nro_comprobante, is_deleted=False).exists():
            raise ValidationError({'nro_comprobante': f'Ya existe un comprobante con el número {nro_comprobante}.'})

        monto_hora   = Decimal(str(data['monto_hora']))
        total_hora   = sum(Decimal(str(b['horas'])) for b in data['bloques'])
        monto_total  = (total_hora * monto_hora).quantize(Decimal('0.01'))
        total_pagado = sum(Decimal(str(v['monto'])) for v in data['valores_pagados'])

        if total_pagado > monto_total:
            raise ValidationError({'valores_pagados': 'El total pagado no puede superar el monto total a pagar.'})

        if total_pagado >= monto_total:
            estado = 'pagado'
            saldo  = Decimal('0')
        elif total_pagado > Decimal('0'):
            estado = 'parcial'
            saldo  = monto_total - total_pagado
        else:
            estado = 'pendiente'
            saldo  = monto_total

        pago = PagoPrestador.objects.create(
            persona_rrhh    = medico,
            nro_comprobante = nro_comprobante,
            fecha_pago      = data['fecha_pago'],
            monto_hora      = monto_hora,
            total_hora      = total_hora,
            monto_total     = monto_total,
            saldo           = saldo,
            estado          = estado,
            id_usu_creator  = request.user,
        )

        for bloque in data['bloques']:
            Agenda.objects.filter(
                id__in=bloque['agenda_ids'],
                is_deleted=False,
            ).update(
                pagado_prestador   = True,
                pago_prestador     = pago,
                id_usu_modificator = request.user,
            )

        for val in data['valores_pagados']:
            try:
                fp  = FormaPago.objects.get(pk=val['forma_pago_id'])
                cta = CuentaMcb.objects.get(pk=val['cta_id'], is_deleted=False)
            except (FormaPago.DoesNotExist, CuentaMcb.DoesNotExist) as e:
                raise ValidationError({'valores_pagados': str(e)})

            det = PagoPrestadorDetCobranza.objects.create(
                pago_prestador = pago,
                forma_pago     = fp,
                cta            = cta,
                monto          = val['monto'],
                voucher        = val.get('voucher', ''),
                id_usu_creator = request.user,
            )

            MovimientoCajaBanco.objects.create(
                cta             = cta,
                fecha           = data['fecha_pago'],
                nro_comprobante = val.get('voucher', '') or None,
                monto_ingreso   = Decimal('0'),
                monto_egreso    = val['monto'],
                vuelto          = Decimal('0'),
                ppdc_id         = det.id,
                id_usu_creator  = request.user,
            )

        self._registrar('CREAR', pago, datos_antes=None, datos_despues=_serializar(pago))
        return Response(PagoPrestadorDetalleSerializer(pago).data, status=201)

    def perform_destroy(self, instance):
        now     = timezone.now()
        det_ids = list(instance.detalle_cobranza.filter(is_deleted=False).values_list('id', flat=True))

        MovimientoCajaBanco.objects.filter(ppdc_id__in=det_ids, is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=self.request.user
        )
        Agenda.objects.filter(pago_prestador=instance, is_deleted=False).update(
            pagado_prestador   = False,
            pago_prestador     = None,
            id_usu_modificator = self.request.user,
        )
        instance.detalle_cobranza.filter(is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=self.request.user,
        )
        super().perform_destroy(instance)

    @action(detail=False, methods=['get'], url_path='siguiente-numero')
    def siguiente_numero(self, request):
        return Response({'siguiente': _siguiente_nro()})

    @action(detail=False, methods=['get'], url_path='medicos-con-pendientes')
    def medicos_con_pendientes(self, request):
        q = request.query_params.get('search', '')
        rrhh_ids = (
            Agenda.objects
            .filter(pagado_prestador=False, is_deleted=False)
            .values_list('horario_prestador__persona_rrhh_id', flat=True)
            .distinct()
        )
        qs = PersonaRRHH.objects.filter(id__in=rrhh_ids, is_deleted=False).select_related('persona')
        if q:
            qs = qs.filter(
                Q(persona__razon_social__icontains=q) | Q(persona__nro_documento__icontains=q)
            )
        return Response([
            {'id': m.id, 'nombre': m.persona.razon_social, 'documento': m.persona.nro_documento}
            for m in qs[:20]
        ])

    @action(detail=False, methods=['get'], url_path='validar-numero')
    def validar_numero(self, request):
        nro = request.query_params.get('nro', '')
        if not nro or not nro.isdigit() or int(nro) < 1:
            return Response({'disponible': False, 'mensaje': 'Número inválido.'})
        existe = PagoPrestador.objects.filter(nro_comprobante=int(nro), is_deleted=False).exists()
        if existe:
            return Response({'disponible': False, 'mensaje': f'El número {nro} ya está registrado.'})
        return Response({'disponible': True, 'mensaje': ''})

    def _qs_con_filtros(self, request):
        qs = PagoPrestador.objects.filter(is_deleted=False).select_related('persona_rrhh__persona')
        if rrhh_id := request.query_params.get('persona_rrhh'):
            qs = qs.filter(persona_rrhh_id=rrhh_id)
        if estado := request.query_params.get('estado'):
            qs = qs.filter(estado=estado)
        if desde := request.query_params.get('fecha_desde'):
            qs = qs.filter(fecha_pago__gte=desde)
        if hasta := request.query_params.get('fecha_hasta'):
            qs = qs.filter(fecha_pago__lte=hasta)
        if search := request.query_params.get('search'):
            qs = qs.filter(
                Q(persona_rrhh__persona__razon_social__icontains=search) |
                Q(persona_rrhh__persona__nro_documento__icontains=search)
            )
        return qs.order_by('-fecha_pago', '-fecha_creacion')

    def _filtros_str(self, request):
        partes = []
        if v := request.query_params.get('search'):
            partes.append(f'Búsqueda: {v}')
        if v := request.query_params.get('estado'):
            partes.append(f'Estado: {dict(PagoPrestador.ESTADO_CHOICES).get(v, v)}')
        if v := request.query_params.get('fecha_desde'):
            partes.append(f'Desde: {v}')
        if v := request.query_params.get('fecha_hasta'):
            partes.append(f'Hasta: {v}')
        return ' · '.join(partes)

    @action(detail=False, methods=['get'], url_path='reporte-pdf')
    def reporte_pdf(self, request):
        ESTADO_LABEL = dict(PagoPrestador.ESTADO_CHOICES)
        filas = [
            {
                'nro':            i,
                'medico':         p.persona_rrhh.persona.razon_social,
                'nro_comprobante': str(p.nro_comprobante).zfill(7) if p.nro_comprobante else '—',
                'fecha_pago':     p.fecha_pago,
                'total_hora':     p.total_hora,
                'monto_hora':     p.monto_hora,
                'monto_total':    p.monto_total,
                'estado':         ESTADO_LABEL.get(p.estado, p.estado),
            }
            for i, p in enumerate(self._qs_con_filtros(request), start=1)
        ]
        hoy = timezone.now().date()
        html = render_to_string('informes/pago_prestador_lista.html', {
            'filas':       filas,
            'fecha':       hoy,
            'total':       len(filas),
            'filtros_str': self._filtros_str(request),
        }, request=request)
        try:
            import weasyprint
            pdf_bytes = weasyprint.HTML(
                string=html, base_url=request.build_absolute_uri('/')
            ).write_pdf()
        except Exception as e:
            return HttpResponse(f'Error generando PDF: {e}', status=500)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="pagos_prestadores_{hoy.strftime("%Y%m%d")}.pdf"'
        return response

    @action(detail=False, methods=['get'], url_path='reporte-excel')
    def reporte_excel(self, request):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse('openpyxl no está instalado.', status=500)

        ESTADO_LABEL = dict(PagoPrestador.ESTADO_CHOICES)
        filas = [
            {
                'nro':            i,
                'medico':         p.persona_rrhh.persona.razon_social,
                'nro_comprobante': str(p.nro_comprobante).zfill(7) if p.nro_comprobante else '—',
                'fecha_pago':     p.fecha_pago.strftime('%d/%m/%Y'),
                'total_hora':     float(p.total_hora),
                'monto_hora':     int(p.monto_hora),
                'monto_total':    int(p.monto_total),
                'estado':         ESTADO_LABEL.get(p.estado, p.estado),
            }
            for i, p in enumerate(self._qs_con_filtros(request), start=1)
        ]
        hoy = timezone.now().date()

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = 'Pagos a Prestadores'

        COLOR_PRIM   = '1A3A5C'
        COLOR_PAR    = 'F8FAFC'
        COLOR_BORDE  = 'E8EDF2'
        fill_header  = PatternFill('solid', fgColor=COLOR_PRIM)
        font_header  = Font(color='FFFFFF', bold=True, size=10)
        font_titulo  = Font(color=COLOR_PRIM, bold=True, size=13)
        font_meta    = Font(color='555555', size=9)
        fill_par     = PatternFill('solid', fgColor=COLOR_PAR)
        thin_border  = Border(bottom=Side(style='thin', color=COLOR_BORDE))
        align_center = Alignment(horizontal='center', vertical='center')
        align_left   = Alignment(horizontal='left',   vertical='center')
        align_right  = Alignment(horizontal='right',  vertical='center')

        NCOLS = 8
        cur   = 1

        ws.merge_cells(f'A{cur}:{get_column_letter(NCOLS)}{cur}')
        c = ws.cell(cur, 1, 'Clínica Lichi — Pagos a Prestadores')
        c.font = font_titulo; c.alignment = align_left
        ws.row_dimensions[cur].height = 20
        cur += 1

        filtros_str = self._filtros_str(request)
        ws.merge_cells(f'A{cur}:{get_column_letter(NCOLS)}{cur}')
        meta_txt = f'Generado el {hoy.strftime("%d/%m/%Y")}  —  {len(filas)} registro{"s" if len(filas) != 1 else ""}'
        if filtros_str:
            meta_txt += f'  —  Filtros: {filtros_str}'
        c = ws.cell(cur, 1, meta_txt)
        c.font = font_meta; c.alignment = align_left
        cur += 2

        cabeceras = ['N°', 'Médico / Prestador', 'Nro. Comp.', 'Fecha pago',
                     'Total horas', 'Monto/hora', 'Monto total', 'Estado']
        for col, txt in enumerate(cabeceras, start=1):
            c = ws.cell(cur, col, txt)
            c.fill = fill_header; c.font = font_header
            c.alignment = align_center if col in (1, 3, 4, 5, 8) else align_left
        ws.row_dimensions[cur].height = 18
        cur += 1

        for f in filas:
            es_par = f['nro'] % 2 == 0
            vals   = [f['nro'], f['medico'], f['nro_comprobante'], f['fecha_pago'],
                      f['total_hora'], f['monto_hora'], f['monto_total'], f['estado']]
            aligns = [align_center, align_left, align_center, align_center,
                      align_right,  align_right, align_right,  align_center]
            for col, (val, aln) in enumerate(zip(vals, aligns), start=1):
                c = ws.cell(cur, col, val)
                if es_par: c.fill = fill_par
                c.border = thin_border; c.alignment = aln
            ws.row_dimensions[cur].height = 15
            cur += 1

        for i, ancho in enumerate([6, 34, 11, 13, 10, 14, 15, 12], start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        from io import BytesIO
        buf = BytesIO()
        wb.save(buf); buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="pagos_prestadores_{hoy.strftime("%Y%m%d")}.xlsx"'
        return response

    @action(detail=True, methods=['get'], url_path='recibo-pdf')
    def recibo_pdf(self, request, pk=None):
        try:
            pago = (
                PagoPrestador.objects
                .select_related('persona_rrhh__persona')
                .get(pk=pk, is_deleted=False)
            )
        except PagoPrestador.DoesNotExist:
            return HttpResponse('Pago no encontrado.', status=404)

        detalle = list(
            pago.detalle_cobranza.filter(is_deleted=False).select_related('forma_pago', 'cta')
        )

        turnos = (
            Agenda.objects
            .filter(pago_prestador=pago, is_deleted=False)
            .select_related('horario_prestador')
            .prefetch_related('horario_prestador__especialidades')
            .order_by('fecha', 'horario_prestador_id')
        )

        bloques = {}
        for t in turnos:
            hp  = t.horario_prestador
            key = (hp.id, str(t.fecha))
            if key not in bloques:
                hora_desde = hp.hora_desde
                hora_hasta = hp.hora_hasta
                segundos = (
                    hora_hasta.hour * 3600 + hora_hasta.minute * 60 + hora_hasta.second
                ) - (
                    hora_desde.hour * 3600 + hora_desde.minute * 60 + hora_desde.second
                )
                horas = round(max(segundos, 0) / 3600, 2)
                especialidades = ', '.join(
                    e.descripcion for e in hp.especialidades.filter(is_deleted=False)
                )
                monto_bloque = int(
                    (Decimal(str(horas)) * pago.monto_hora).quantize(Decimal('1'))
                )
                bloques[key] = {
                    'fecha':       t.fecha,
                    'hora_desde':  str(hora_desde)[:5],
                    'hora_hasta':  str(hora_hasta)[:5],
                    'horas':       horas,
                    'especialidad': especialidades,
                    'monto_bloque': monto_bloque,
                }

        nro_fmt = str(pago.nro_comprobante).zfill(7) if pago.nro_comprobante else '—'

        html = render_to_string('informes/recibo_prestador.html', {
            'pago':    pago,
            'nro_fmt': nro_fmt,
            'bloques': list(bloques.values()),
            'detalle': detalle,
        }, request=request)

        try:
            import weasyprint
            pdf_bytes = weasyprint.HTML(
                string=html, base_url=request.build_absolute_uri('/')
            ).write_pdf()
        except Exception as e:
            return HttpResponse(f'Error generando PDF: {e}', status=500)

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="recibo_{nro_fmt}.pdf"'
        return response

    @action(detail=False, methods=['get'], url_path='bloques-pendientes')
    def bloques_pendientes(self, request):
        rrhh_id    = request.query_params.get('persona_rrhh')
        fecha_hasta = request.query_params.get('fecha_hasta')

        if not rrhh_id:
            return Response({'error': 'Se requiere persona_rrhh.'}, status=400)

        filtros = {
            'horario_prestador__persona_rrhh_id': rrhh_id,
            'pagado_prestador': False,
            'is_deleted': False,
        }
        if fecha_hasta:
            filtros['fecha__lte'] = fecha_hasta

        turnos = (
            Agenda.objects
            .filter(**filtros)
            .select_related('horario_prestador__persona_rrhh')
            .prefetch_related('horario_prestador__especialidades')
            .order_by('fecha', 'horario_prestador_id')
        )

        bloques = {}
        for t in turnos:
            hp  = t.horario_prestador
            key = (hp.id, str(t.fecha))
            if key not in bloques:
                hora_desde = hp.hora_desde
                hora_hasta = hp.hora_hasta
                segundos   = (
                    hora_hasta.hour * 3600 + hora_hasta.minute * 60 + hora_hasta.second
                ) - (
                    hora_desde.hour * 3600 + hora_desde.minute * 60 + hora_desde.second
                )
                horas = round(max(segundos, 0) / 3600, 2)
                especialidades = ', '.join(e.descripcion for e in hp.especialidades.filter(is_deleted=False))
                bloques[key] = {
                    'horario_prestador_id': hp.id,
                    'fecha':       str(t.fecha),
                    'hora_desde':  str(hora_desde)[:5],
                    'hora_hasta':  str(hora_hasta)[:5],
                    'horas':       horas,
                    'especialidad': especialidades,
                    'agenda_ids':  [],
                }
            bloques[key]['agenda_ids'].append(t.id)

        return Response(list(bloques.values()))
