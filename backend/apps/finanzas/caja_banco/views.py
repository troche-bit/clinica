from datetime import timedelta
from decimal import Decimal

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from django.db.models import Sum, Count, Q
from django.db.models.functions import Coalesce

from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError

from apps.administracion.auditoria.mixins import AuditoriaMixin
from apps.core.permissions import IsAdminRole, IsAdminOrRecepcionista
from .models import CuentaMcb, MovimientoCajaBanco
from .serializers import (
    CuentaMcbListSerializer,
    CuentaMcbSerializer,
    MovimientoCajaBancoListSerializer,
    MovimientoCajaBancoSerializer,
)


class CuentaMcbViewSet(AuditoriaMixin, viewsets.ModelViewSet):
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields   = ['descripcion']
    ordering_fields = ['descripcion']
    ordering        = ['descripcion']

    def get_permissions(self):
        if self.action in ('list', 'retrieve', 'eliminados', 'dashboard_mensual'):
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsAdminRole()]

    def get_serializer_class(self):
        if self.action in ('list', 'retrieve', 'eliminados'):
            return CuentaMcbListSerializer
        return CuentaMcbSerializer

    def _qs_anotado(self, deleted=False):
        return CuentaMcb.objects.filter(is_deleted=deleted).annotate(
            saldo=Coalesce(
                Sum('movimientos__monto_ingreso', filter=Q(movimientos__is_deleted=False)),
                Decimal('0'),
            ) - Coalesce(
                Sum('movimientos__monto_egreso', filter=Q(movimientos__is_deleted=False)),
                Decimal('0'),
            ),
            total_movimientos=Count('movimientos', filter=Q(movimientos__is_deleted=False)),
        )

    def get_queryset(self):
        return self._qs_anotado(deleted=False)

    def perform_destroy(self, instance):
        if instance.movimientos.filter(is_deleted=False).exists():
            raise ValidationError('No se puede eliminar una cuenta con movimientos activos.')
        super().perform_destroy(instance)

    @action(detail=False, methods=['get'], url_path='eliminados')
    def eliminados(self, request):
        serializer = self.get_serializer(self._qs_anotado(deleted=True), many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='dashboard-mensual')
    def dashboard_mensual(self, request):
        MESES = ['enero','febrero','marzo','abril','mayo','junio','julio',
                 'agosto','septiembre','octubre','noviembre','diciembre']
        hoy        = timezone.localtime().date()
        inicio_mes = hoy.replace(day=1)

        cuentas_qs = self._qs_anotado(deleted=False).order_by('descripcion')
        cuentas = [
            {'id': c.id, 'descripcion': c.descripcion, 'saldo': float(c.saldo)}
            for c in cuentas_qs
        ]

        movs_mes = MovimientoCajaBanco.objects.filter(
            is_deleted=False, fecha__gte=inicio_mes, fecha__lte=hoy,
        )
        agg = movs_mes.aggregate(ing=Sum('monto_ingreso'), egr=Sum('monto_egreso'))
        total_ingresos = float(agg['ing'] or 0)
        total_egresos  = float(agg['egr'] or 0)

        por_dia_qs = movs_mes.values('fecha').annotate(
            ingresos=Sum('monto_ingreso'),
            egresos=Sum('monto_egreso'),
        ).order_by('fecha')
        por_dia_map = {
            item['fecha'].isoformat(): {
                'ingresos': float(item['ingresos'] or 0),
                'egresos':  float(item['egresos']  or 0),
            }
            for item in por_dia_qs
        }

        dias = []
        d = inicio_mes
        while d <= hoy:
            k = d.isoformat()
            dias.append({
                'fecha':    k,
                'dia':      d.day,
                'ingresos': por_dia_map.get(k, {}).get('ingresos', 0),
                'egresos':  por_dia_map.get(k, {}).get('egresos',  0),
            })
            d += timedelta(days=1)

        return Response({
            'total_ingresos_mes': total_ingresos,
            'total_egresos_mes':  total_egresos,
            'saldo_neto_mes':     total_ingresos - total_egresos,
            'cuentas':            cuentas,
            'por_dia':            dias,
            'mes':                f"{MESES[inicio_mes.month - 1]} {inicio_mes.year}",
            'fecha':              hoy.isoformat(),
        })


class MovimientoCajaBancoViewSet(AuditoriaMixin, viewsets.ModelViewSet):
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields   = ['nro_comprobante']
    ordering_fields = ['fecha', 'monto_ingreso', 'monto_egreso']
    ordering        = ['-fecha', '-fecha_creacion']

    def get_permissions(self):
        if self.action in ('list', 'retrieve', 'eliminados', 'reporte_pdf', 'reporte_excel'):
            return [IsAuthenticated()]
        if self.action == 'destroy':
            return [IsAuthenticated(), IsAdminRole()]
        return [IsAuthenticated(), IsAdminOrRecepcionista()]

    def get_serializer_class(self):
        if self.action in ('list', 'retrieve', 'eliminados'):
            return MovimientoCajaBancoListSerializer
        return MovimientoCajaBancoSerializer

    def get_queryset(self):
        qs = MovimientoCajaBanco.objects.filter(is_deleted=False).select_related('cta')

        cta_id = self.request.query_params.get('cta')
        if cta_id:
            qs = qs.filter(cta_id=cta_id)

        tipo = self.request.query_params.get('tipo')
        if tipo == 'ingreso':
            qs = qs.filter(monto_ingreso__gt=0)
        elif tipo == 'egreso':
            qs = qs.filter(monto_egreso__gt=0)

        fecha_desde = self.request.query_params.get('fecha_desde')
        fecha_hasta = self.request.query_params.get('fecha_hasta')
        if fecha_desde:
            qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha__lte=fecha_hasta)

        return qs

    def perform_destroy(self, instance):
        if instance.vfdc_id or instance.vrc_id or instance.ppdc_id:
            raise ValidationError('No se puede eliminar un movimiento generado automáticamente.')
        super().perform_destroy(instance)

    @action(detail=False, methods=['get'], url_path='eliminados')
    def eliminados(self, request):
        qs = MovimientoCajaBanco.objects.filter(is_deleted=True).select_related('cta')
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    def _qs_con_filtros(self, request):
        qs = MovimientoCajaBanco.objects.filter(is_deleted=False).select_related('cta')
        if cta_id := request.query_params.get('cta'):
            qs = qs.filter(cta_id=cta_id)
        tipo = request.query_params.get('tipo')
        if tipo == 'ingreso':
            qs = qs.filter(monto_ingreso__gt=0)
        elif tipo == 'egreso':
            qs = qs.filter(monto_egreso__gt=0)
        if desde := request.query_params.get('fecha_desde'):
            qs = qs.filter(fecha__gte=desde)
        if hasta := request.query_params.get('fecha_hasta'):
            qs = qs.filter(fecha__lte=hasta)
        return qs.order_by('-fecha', '-fecha_creacion')

    def _filtros_str(self, request):
        partes = []
        if cta_id := request.query_params.get('cta'):
            try:
                partes.append(f'Cuenta: {CuentaMcb.objects.get(pk=cta_id).descripcion}')
            except CuentaMcb.DoesNotExist:
                pass
        tipo = request.query_params.get('tipo')
        if tipo == 'ingreso':
            partes.append('Tipo: Ingresos')
        elif tipo == 'egreso':
            partes.append('Tipo: Egresos')
        if v := request.query_params.get('fecha_desde'):
            partes.append(f'Desde: {v}')
        if v := request.query_params.get('fecha_hasta'):
            partes.append(f'Hasta: {v}')
        return ' · '.join(partes)

    @action(detail=False, methods=['get'], url_path='reporte-pdf')
    def reporte_pdf(self, request):
        import weasyprint
        qs    = self._qs_con_filtros(request)
        filas = list(qs)
        hoy   = timezone.localtime().date()
        total_ingresos = sum(m.monto_ingreso for m in filas)
        total_egresos  = sum(m.monto_egreso  for m in filas)
        ctx = {
            'filas': [
                {
                    'nro':            i,
                    'fecha':          m.fecha,
                    'cuenta':         m.cta.descripcion,
                    'nro_comprobante': m.nro_comprobante or '—',
                    'ingreso':        m.monto_ingreso,
                    'egreso':         m.monto_egreso,
                }
                for i, m in enumerate(filas, 1)
            ],
            'total':          len(filas),
            'total_ingresos': total_ingresos,
            'total_egresos':  total_egresos,
            'fecha':          hoy,
            'filtros_str':    self._filtros_str(request),
        }
        try:
            html = render_to_string('informes/movimiento_lista.html', ctx, request=request)
            pdf  = weasyprint.HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
        except Exception as e:
            return HttpResponse(f'Error generando PDF: {e}', status=500)
        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = f'inline; filename="movimientos_{hoy.strftime("%Y%m%d")}.pdf"'
        return resp

    @action(detail=False, methods=['get'], url_path='reporte-excel')
    def reporte_excel(self, request):
        import openpyxl
        from io import BytesIO
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        qs    = self._qs_con_filtros(request)
        filas = list(qs)
        hoy   = timezone.localtime().date()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Movimientos'

        fill_hdr  = PatternFill('solid', fgColor='1A3A5C')
        font_hdr  = Font(color='FFFFFF', bold=True, size=10)
        font_tit  = Font(color='1A3A5C', bold=True, size=13)
        font_meta = Font(color='555555', size=9)
        fill_par  = PatternFill('solid', fgColor='F8FAFC')
        thin      = Border(bottom=Side(style='thin', color='E8EDF2'))
        align_c   = Alignment(horizontal='center', vertical='center')
        align_l   = Alignment(horizontal='left',   vertical='center')
        align_r   = Alignment(horizontal='right',  vertical='center')

        NCOLS = 6
        cur   = 1
        ws.merge_cells(f'A{cur}:{get_column_letter(NCOLS)}{cur}')
        c = ws.cell(cur, 1, 'Clínica Lichi — Movimientos de Caja/Banco')
        c.font = font_tit; c.alignment = align_l
        ws.row_dimensions[cur].height = 20
        cur += 1

        filtros_str = self._filtros_str(request)
        ws.merge_cells(f'A{cur}:{get_column_letter(NCOLS)}{cur}')
        meta = f'Generado el {hoy.strftime("%d/%m/%Y")}  —  {len(filas)} registro{"s" if len(filas) != 1 else ""}'
        if filtros_str:
            meta += f'  —  Filtros: {filtros_str}'
        c = ws.cell(cur, 1, meta)
        c.font = font_meta; c.alignment = align_l
        cur += 2

        for col, txt in enumerate(['N°', 'Fecha', 'Cuenta', 'N° Comprobante', 'Ingreso (Gs.)', 'Egreso (Gs.)'], 1):
            c = ws.cell(cur, col, txt)
            c.fill = fill_hdr; c.font = font_hdr
            c.alignment = align_c if col in (1, 2, 4) else align_l
        ws.row_dimensions[cur].height = 18
        cur += 1

        for i, m in enumerate(filas, 1):
            vals   = [i, m.fecha.strftime('%d/%m/%Y'), m.cta.descripcion, m.nro_comprobante or '—', int(m.monto_ingreso), int(m.monto_egreso)]
            aligns = [align_c, align_c, align_l, align_c, align_r, align_r]
            for col, (val, aln) in enumerate(zip(vals, aligns), 1):
                c = ws.cell(cur, col, val)
                if i % 2 == 0: c.fill = fill_par
                c.border = thin; c.alignment = aln
            ws.row_dimensions[cur].height = 15
            cur += 1

        for i, ancho in enumerate([5, 12, 30, 16, 16, 16], 1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        buf = BytesIO()
        wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="movimientos_{hoy.strftime("%Y%m%d")}.xlsx"'
        return resp
