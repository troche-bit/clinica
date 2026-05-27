from decimal import Decimal
from django.db import transaction
from django.db.models import Max, Q
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone

from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.administracion.auditoria.mixins import AuditoriaMixin
from apps.core.permissions import IsAdminRole, IsAdminOrRecepcionista
from apps.administracion.persona.models import Persona
from apps.forma_pago.models import FormaPago
from apps.finanzas.caja_banco.models import CuentaMcb, MovimientoCajaBanco
from apps.finanzas.estadocuenta.models import CtaCobrar

from .models import Cobranza, CobranzaDet, ValorRecibidoCob
from .serializers import (
    CobranzaListSerializer,
    CobranzaDetalleSerializer,
    CobranzaCreateSerializer,
)


def _siguiente_comprobante_nro():
    max_nro = Cobranza.objects.filter(is_deleted=False).aggregate(
        max_nro=Max('comprobante_nro')
    )['max_nro']
    return (max_nro or 0) + 1


class CobranzaViewSet(AuditoriaMixin, viewsets.ModelViewSet):
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields   = ['persona__razon_social', 'persona__nro_documento']
    ordering_fields = ['fecha', 'monto', 'fecha_creacion']
    ordering        = ['-fecha', '-fecha_creacion']

    def get_permissions(self):
        if self.action in ('list', 'retrieve', 'siguiente_numero', 'validar_numero',
                           'cuotas_pendientes', 'clientes_con_pendientes',
                           'recibo_pdf', 'reporte_pdf', 'reporte_excel'):
            return [IsAuthenticated()]
        if self.action in ('destroy', 'eliminados'):
            return [IsAuthenticated(), IsAdminRole()]
        return [IsAuthenticated(), IsAdminOrRecepcionista()]

    def _qs_con_filtros(self, request):
        qs = Cobranza.objects.filter(is_deleted=False).select_related('persona')
        search      = request.query_params.get('search', '')
        fecha_desde = request.query_params.get('fecha_desde', '')
        fecha_hasta = request.query_params.get('fecha_hasta', '')
        if search:
            qs = qs.filter(
                Q(persona__razon_social__icontains=search) |
                Q(persona__nro_documento__icontains=search)
            )
        if fecha_desde:
            qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha__lte=fecha_hasta)
        return qs.order_by('-fecha', '-fecha_creacion')

    def _filtros_str(self, request):
        partes = []
        if request.query_params.get('search'):
            partes.append(f'Búsqueda: "{request.query_params["search"]}"')
        if request.query_params.get('fecha_desde'):
            partes.append(f'Desde: {request.query_params["fecha_desde"]}')
        if request.query_params.get('fecha_hasta'):
            partes.append(f'Hasta: {request.query_params["fecha_hasta"]}')
        return ' · '.join(partes) if partes else ''

    def get_queryset(self):
        qs = Cobranza.objects.filter(is_deleted=False).select_related('persona')

        fecha_desde = self.request.query_params.get('fecha_desde')
        fecha_hasta = self.request.query_params.get('fecha_hasta')
        if fecha_desde:
            qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha__lte=fecha_hasta)

        return qs

    def get_serializer_class(self):
        if self.action == 'create':
            return CobranzaCreateSerializer
        if self.action == 'retrieve':
            return CobranzaDetalleSerializer
        return CobranzaListSerializer

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        ser = CobranzaCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        try:
            persona = Persona.objects.get(pk=data['persona'], is_deleted=False)
        except Persona.DoesNotExist:
            raise ValidationError({'persona': 'Persona no encontrada.'})

        comprobante_nro = data.get('comprobante_nro') or _siguiente_comprobante_nro()
        if Cobranza.objects.filter(comprobante_nro=comprobante_nro, is_deleted=False).exists():
            raise ValidationError({'comprobante_nro': f'El número {comprobante_nro} ya está en uso.'})

        cuotas_data = []
        monto_total = Decimal('0')
        for det in data['detalle']:
            try:
                cta = CtaCobrar.objects.select_for_update().get(pk=det['cta_cobrar_id'], is_deleted=False)
            except CtaCobrar.DoesNotExist:
                raise ValidationError({'detalle': f"Cuota ID {det['cta_cobrar_id']} no encontrada."})

            if det['monto_pagado'] <= Decimal('0'):
                raise ValidationError({'detalle': f'El monto a pagar de la cuota {cta.id} debe ser mayor a 0.'})

            if det['monto_pagado'] > cta.saldo:
                raise ValidationError({
                    'detalle': f"El monto a pagar ({det['monto_pagado']}) supera el saldo disponible ({cta.saldo}) de la cuota {cta.id}."
                })

            cuotas_data.append({'cta': cta, 'det': det})
            monto_total += Decimal(str(det['monto_pagado']))

        total_recibido = sum(Decimal(str(v['monto'])) for v in data['valores_recibidos'])
        vuelto         = max(Decimal('0'), total_recibido - monto_total)

        cab = Cobranza.objects.create(
            fecha           = data['fecha'],
            persona         = persona,
            comprobante_nro = comprobante_nro,
            monto           = monto_total,
            vuelto          = vuelto,
            id_usu_creator  = request.user,
        )

        for item in cuotas_data:
            cta            = item['cta']
            det            = item['det']
            saldo_anterior = cta.saldo

            CobranzaDet.objects.create(
                cobranza        = cab,
                cta_cobrar      = cta,
                monto_total     = saldo_anterior,
                monto_pagado    = det['monto_pagado'],
                nro_comprobante = det.get('nro_comprobante', ''),
                id_usu_creator  = request.user,
            )

            nuevo_saldo = saldo_anterior - Decimal(str(det['monto_pagado']))
            cta.saldo   = nuevo_saldo
            if nuevo_saldo <= Decimal('0'):
                cta.estado = 'pagado'
            cta.id_usu_modificator = request.user
            cta.save()

        # Distribuye el ingreso FIFO: solo hasta cubrir monto_total.
        # ValorRecibidoCob registra lo recibido realmente; MovimientoCajaBanco
        # solo registra el monto que efectivamente ingresa a la cuenta.
        saldo_pendiente = monto_total
        for val in data['valores_recibidos']:
            try:
                fp  = FormaPago.objects.get(pk=val['forma_pago_id'])
                cta = CuentaMcb.objects.get(pk=val['cta_id'], is_deleted=False)
            except (FormaPago.DoesNotExist, CuentaMcb.DoesNotExist) as e:
                raise ValidationError({'valores_recibidos': str(e)})

            vrc = ValorRecibidoCob.objects.create(
                cobranza        = cab,
                forma_pago      = fp,
                cta             = cta,
                monto           = val['monto'],
                voucher         = val.get('voucher', ''),
                nro_comprobante = val.get('nro_comprobante', ''),
                id_usu_creator  = request.user,
            )

            ingreso = min(Decimal(str(val['monto'])), saldo_pendiente)
            if ingreso > Decimal('0'):
                MovimientoCajaBanco.objects.create(
                    cta             = cta,
                    fecha           = data['fecha'],
                    nro_comprobante = val.get('nro_comprobante', '') or None,
                    monto_ingreso   = ingreso,
                    monto_egreso    = Decimal('0'),
                    vuelto          = Decimal('0'),
                    vrc_id          = vrc.id,
                    id_usu_creator  = request.user,
                )
                saldo_pendiente -= ingreso

        return Response(CobranzaDetalleSerializer(cab).data, status=201)

    def perform_destroy(self, instance):
        now     = timezone.now()
        hoy     = timezone.localtime(now).date()
        vrc_ids = list(instance.valores_recibidos.filter(is_deleted=False).values_list('id', flat=True))

        for det in instance.detalle.filter(is_deleted=False).select_related('cta_cobrar'):
            cta        = det.cta_cobrar
            cta.saldo += det.monto_pagado
            cta.estado = 'vencido' if cta.fecha_vencimiento < hoy else 'pendiente'
            cta.id_usu_modificator = self.request.user
            cta.save()

        MovimientoCajaBanco.objects.filter(vrc_id__in=vrc_ids, is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=self.request.user
        )
        instance.valores_recibidos.filter(is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=self.request.user
        )
        instance.detalle.filter(is_deleted=False).update(
            is_deleted=True, fecha_eliminacion=now, id_usu_modificator=self.request.user
        )
        super().perform_destroy(instance)

    @action(detail=False, methods=['get'], url_path='siguiente-numero')
    def siguiente_numero(self, request):
        return Response({'siguiente': _siguiente_comprobante_nro()})

    @action(detail=False, methods=['get'], url_path='validar-numero')
    def validar_numero(self, request):
        nro = request.query_params.get('nro', '').strip()
        if not nro or not nro.isdigit():
            return Response({'disponible': False, 'mensaje': 'Número inválido.'}, status=400)
        existe = Cobranza.objects.filter(comprobante_nro=int(nro), is_deleted=False).exists()
        return Response({
            'disponible': not existe,
            'mensaje': '' if not existe else f'El número {nro} ya está en uso.',
        })

    @action(detail=False, methods=['get'], url_path='clientes-con-pendientes')
    def clientes_con_pendientes(self, request):
        from apps.administracion.persona.models import Persona as PersonaModel
        search = request.query_params.get('search', '').strip()
        qs = PersonaModel.objects.filter(
            is_deleted=False,
            facturas__is_deleted=False,
            facturas__cuotas__is_deleted=False,
            facturas__cuotas__saldo__gt=0,
        ).distinct()
        if search:
            qs = qs.filter(
                Q(razon_social__icontains=search) | Q(nro_documento__icontains=search)
            )
        resultado = [
            {'id': p.id, 'razon_social': p.razon_social, 'nro_documento': p.nro_documento}
            for p in qs.order_by('razon_social')[:20]
        ]
        return Response(resultado)

    @action(detail=False, methods=['get'], url_path='eliminados')
    def eliminados(self, request):
        qs = Cobranza.objects.filter(is_deleted=True).select_related('persona').order_by('-fecha_eliminacion')
        return Response(CobranzaListSerializer(qs, many=True).data)

    @action(detail=True, methods=['get'], url_path='recibo-pdf')
    def recibo_pdf(self, request, pk=None):
        import weasyprint
        cobranza = Cobranza.objects.select_related('persona').prefetch_related(
            'detalle__cta_cobrar__vfc__timbrado',
            'valores_recibidos__forma_pago',
            'valores_recibidos__cta',
        ).get(pk=pk, is_deleted=False)

        nro_fmt = str(cobranza.comprobante_nro).zfill(7)

        detalle_rows = []
        for det in cobranza.detalle.filter(is_deleted=False).select_related(
            'cta_cobrar__vfc__timbrado'
        ):
            vfc   = det.cta_cobrar.vfc
            estab = vfc.establecimiento if vfc.establecimiento else str(vfc.timbrado.punto_sucursal).zfill(3)
            expd  = vfc.expedicion if vfc.expedicion else str(vfc.timbrado.punto_expedicion).zfill(3)
            detalle_rows.append({
                'factura_nro':       f'{estab}-{expd}-{str(vfc.nro_comprobante).zfill(7)}',
                'cuota_display':     f'{det.cta_cobrar.nro_cuota}/{det.cta_cobrar.cant_cuota}',
                'fecha_vencimiento': det.cta_cobrar.fecha_vencimiento,
                'monto_total':       det.monto_total,
                'monto_pagado':      det.monto_pagado,
            })

        valores = cobranza.valores_recibidos.filter(is_deleted=False).select_related('forma_pago', 'cta')

        ctx = {
            'cobranza':  cobranza,
            'nro_fmt':   nro_fmt,
            'detalle':   detalle_rows,
            'valores':   valores,
        }
        try:
            html_str  = render_to_string('informes/recibo_cobranza.html', ctx, request=request)
            pdf_bytes = weasyprint.HTML(
                string=html_str,
                base_url=request.build_absolute_uri('/'),
            ).write_pdf()
            resp = HttpResponse(pdf_bytes, content_type='application/pdf')
            resp['Content-Disposition'] = f'inline; filename="recibo_cob_{nro_fmt}.pdf"'
            return resp
        except Exception as e:
            return Response({'error': f'Error generando PDF: {e}'}, status=500)

    @action(detail=False, methods=['get'], url_path='reporte-pdf')
    def reporte_pdf(self, request):
        import weasyprint
        qs    = self._qs_con_filtros(request)
        filas = []
        for i, c in enumerate(qs, start=1):
            filas.append({
                'nro':         i,
                'comprobante': str(c.comprobante_nro).zfill(7) if c.comprobante_nro else '—',
                'fecha':       c.fecha,
                'cliente':     c.persona.razon_social,
                'documento':   c.persona.nro_documento,
                'monto':       c.monto,
                'vuelto':      c.vuelto,
            })
        ctx = {
            'filas':       filas,
            'total':       len(filas),
            'fecha':       timezone.localtime(timezone.now()).date(),
            'filtros_str': self._filtros_str(request),
        }
        try:
            html_str  = render_to_string('informes/cobranza_lista.html', ctx, request=request)
            pdf_bytes = weasyprint.HTML(
                string=html_str,
                base_url=request.build_absolute_uri('/'),
            ).write_pdf()
            resp = HttpResponse(pdf_bytes, content_type='application/pdf')
            resp['Content-Disposition'] = 'inline; filename="cobranzas_lista.pdf"'
            return resp
        except Exception as e:
            return Response({'error': f'Error generando PDF: {e}'}, status=500)

    @action(detail=False, methods=['get'], url_path='reporte-excel')
    def reporte_excel(self, request):
        import openpyxl
        from io import BytesIO
        from openpyxl.styles import Font, PatternFill, Alignment

        qs = self._qs_con_filtros(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cobranzas'

        hdr_font  = Font(bold=True, color='FFFFFF', size=10)
        hdr_fill  = PatternFill('solid', fgColor='1A3A5C')
        hdr_align = Alignment(horizontal='center', vertical='center')
        alt_fill  = PatternFill('solid', fgColor='F8FAFC')

        cols   = ['#', 'Nro. Comp.', 'Fecha', 'Cliente', 'Documento', 'Monto (Gs.)', 'Vuelto (Gs.)']
        widths = [5, 12, 12, 40, 20, 18, 14]

        for col_idx, (title, width) in enumerate(zip(cols, widths), start=1):
            cell            = ws.cell(row=1, column=col_idx, value=title)
            cell.font       = hdr_font
            cell.fill       = hdr_fill
            cell.alignment  = hdr_align
            ws.column_dimensions[cell.column_letter].width = width

        for i, c in enumerate(qs, start=1):
            row_data = [
                i,
                str(c.comprobante_nro).zfill(7) if c.comprobante_nro else '—',
                c.fecha.strftime('%d/%m/%Y'),
                c.persona.razon_social,
                c.persona.nro_documento,
                int(c.monto),
                int(c.vuelto),
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i + 1, column=col_idx, value=value)
                if i % 2 == 0:
                    cell.fill = alt_fill

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        fecha_str = timezone.localtime(timezone.now()).strftime('%Y%m%d')
        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="cobranzas_{fecha_str}.xlsx"'
        return resp

    @action(detail=False, methods=['get'], url_path='cuotas-pendientes')
    def cuotas_pendientes(self, request):
        persona_id = request.query_params.get('persona')
        if not persona_id:
            return Response({'error': 'Se requiere el parámetro persona.'}, status=400)

        cuotas = CtaCobrar.objects.filter(
            is_deleted=False,
            saldo__gt=0,
            vfc__persona_id=persona_id,
            vfc__is_deleted=False,
        ).select_related('vfc', 'vfc__timbrado').order_by('fecha_vencimiento', 'id')

        resultado = []
        for c in cuotas:
            vfc   = c.vfc
            estab = vfc.establecimiento if vfc.establecimiento else str(vfc.timbrado.punto_sucursal).zfill(3)
            expd  = vfc.expedicion if vfc.expedicion else str(vfc.timbrado.punto_expedicion).zfill(3)
            resultado.append({
                'id':                c.id,
                'nro_cuota':         c.nro_cuota,
                'cant_cuota':        c.cant_cuota,
                'monto_cuota':       str(c.monto_cuota),
                'saldo':             str(c.saldo),
                'fecha_vencimiento': str(c.fecha_vencimiento),
                'estado':            c.estado,
                'factura_nro':       f'{estab}-{expd}-{str(vfc.nro_comprobante).zfill(7)}',
                'factura_fecha':     str(vfc.fecha),
                'factura_id':        vfc.id,
            })

        return Response(resultado)
