from datetime import date
import calendar
from django.template.loader import render_to_string
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count
from django.db.models.functions import TruncDate, TruncMonth
from rest_framework import viewsets, filters
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from apps.administracion.auditoria.mixins import AuditoriaMixin
from .models import Paciente
from .serializers import PacienteSerializer, PacienteListSerializer
from config.pagination import StandardPagination

MESES      = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
               "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
MESES_CORTO = ["Ene","Feb","Mar","Abr","May","Jun",
                "Jul","Ago","Sep","Oct","Nov","Dic"]
ETIQUETAS_SEXO = {"M": "Masculino", "F": "Femenino", "O": "Otro"}


class PacienteViewSet(AuditoriaMixin, viewsets.ModelViewSet):
    pagination_class   = StandardPagination
    permission_classes = [IsAuthenticated]
    filter_backends    = [filters.SearchFilter, filters.OrderingFilter]
    search_fields      = ["persona__razon_social", "persona__nro_documento"]
    ordering_fields    = ["persona__razon_social", "fecha_creacion"]

    def get_queryset(self):
        return Paciente.objects.filter(
            is_deleted=False
        ).select_related(
            "persona__tipo_documento",
            "persona__pais",
            "persona__departamento",
            "persona__ciudad",
            "responsable__persona",
        )

    def get_serializer_class(self):
        if self.action in ["list", "retrieve", "eliminados"]:
            return PacienteListSerializer
        return PacienteSerializer

    def perform_destroy(self, instance):
        from apps.principal.agenda.models import Agenda
        tiene_citas = Agenda.objects.filter(
            paciente=instance,
            is_deleted=False,
            estado__in=[
                Agenda.Estado.DISPONIBLE,
                Agenda.Estado.OCUPADO,
                Agenda.Estado.REALIZADO,
            ],
        ).exists()
        if tiene_citas:
            raise ValidationError(
                'No se puede eliminar un paciente con citas activas (disponible, ocupado o realizado). '
                'Primero cancelá o finalizá todas sus citas.'
            )
        super().perform_destroy(instance)

    @action(detail=False, methods=["get"], url_path="eliminados")
    def eliminados(self, request):
        qs = Paciente.objects.filter(is_deleted=True).select_related("persona")
        serializer = PacienteListSerializer(qs, many=True)
        return Response(serializer.data)

    # ── Helpers para reportes ──────────────────────────────────────

    def _get_queryset_filtrado(self, request):
        qs = (
            Paciente.objects
            .filter(is_deleted=False)
            .select_related("persona__pais", "persona__departamento", "persona__ciudad", "responsable__persona")
            .order_by("persona__razon_social")
        )
        p = request.query_params
        if p.get("sexo"):
            qs = qs.filter(sexo=p["sexo"])
        if p.get("grupo_sanguineo"):
            qs = qs.filter(grupo_sanguineo__iexact=p["grupo_sanguineo"])
        if p.get("pais"):
            qs = qs.filter(persona__pais_id=p["pais"])
        if p.get("departamento"):
            qs = qs.filter(persona__departamento_id=p["departamento"])
        if p.get("ciudad"):
            qs = qs.filter(persona__ciudad_id=p["ciudad"])
        if p.get("fecha_desde"):
            qs = qs.filter(fecha_creacion__date__gte=p["fecha_desde"])
        if p.get("fecha_hasta"):
            qs = qs.filter(fecha_creacion__date__lte=p["fecha_hasta"])
        return qs

    def _build_filtros_aplicados(self, request):
        from apps.mantenimiento.ubicacion.models import Pais, Departamento, Ciudad
        p = request.query_params
        filtros = []
        if p.get("sexo"):
            filtros.append(("Sexo", ETIQUETAS_SEXO.get(p["sexo"], p["sexo"])))
        if p.get("grupo_sanguineo"):
            filtros.append(("Tipo de sangre", p["grupo_sanguineo"].upper()))
        if p.get("pais"):
            try:
                filtros.append(("País", Pais.objects.get(pk=p["pais"]).descripcion))
            except Pais.DoesNotExist:
                pass
        if p.get("departamento"):
            try:
                filtros.append(("Departamento", Departamento.objects.get(pk=p["departamento"]).descripcion))
            except Departamento.DoesNotExist:
                pass
        if p.get("ciudad"):
            try:
                filtros.append(("Ciudad", Ciudad.objects.get(pk=p["ciudad"]).descripcion))
            except Ciudad.DoesNotExist:
                pass
        if p.get("fecha_desde"):
            filtros.append(("Registro desde", p["fecha_desde"]))
        if p.get("fecha_hasta"):
            filtros.append(("Registro hasta", p["fecha_hasta"]))
        return filtros

    def _build_filas(self, pacientes):
        hoy = date.today()

        def calcular_edad(fecha_nac):
            if not fecha_nac:
                return "—"
            return hoy.year - fecha_nac.year - (
                (hoy.month, hoy.day) < (fecha_nac.month, fecha_nac.day)
            )

        return [
            {
                "nro":         i,
                "nombre":      p.persona.razon_social,
                "documento":   p.persona.nro_documento,
                "edad":        calcular_edad(p.fecha_nacimiento),
                "sexo":        p.get_sexo_display(),
                "telefono":    p.persona.telefono or "—",
                "responsable": p.responsable.persona.razon_social if p.responsable else "—",
            }
            for i, p in enumerate(pacientes, start=1)
        ]

    # ── Reportes ──────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="reporte-lista",
            permission_classes=[IsAuthenticated])
    def reporte_lista(self, request):
        pacientes         = self._get_queryset_filtrado(request)
        filtros_aplicados = self._build_filtros_aplicados(request)
        filas             = self._build_filas(pacientes)
        hoy               = date.today()

        contexto = {
            "filas":             filas,
            "fecha":             hoy,
            "total":             len(filas),
            "filtros_aplicados": filtros_aplicados,
        }
        html = render_to_string("informes/paciente_lista.html", contexto, request=request)

        try:
            import weasyprint
            pdf_bytes = weasyprint.HTML(
                string=html, base_url=request.build_absolute_uri("/")
            ).write_pdf()
        except Exception as e:
            return HttpResponse(f"Error generando PDF: {e}", status=500)

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="listado_pacientes.pdf"'
        return response

    @action(detail=False, methods=["get"], url_path="reporte-lista-excel",
            permission_classes=[IsAuthenticated])
    def reporte_lista_excel(self, request):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse("openpyxl no está instalado.", status=500)

        pacientes         = self._get_queryset_filtrado(request)
        filtros_aplicados = self._build_filtros_aplicados(request)
        filas             = self._build_filas(pacientes)
        hoy               = date.today()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listado de Pacientes"

        COLOR_PRIMARIO = "1A3A5C"
        COLOR_FILA_PAR = "F8FAFC"
        COLOR_BORDE    = "E8EDF2"

        fill_header = PatternFill("solid", fgColor=COLOR_PRIMARIO)
        font_header = Font(color="FFFFFF", bold=True, size=10)
        font_titulo = Font(color=COLOR_PRIMARIO, bold=True, size=13)
        font_meta   = Font(color="555555", size=9)
        font_filtro = Font(color="374151", size=9)
        fill_par    = PatternFill("solid", fgColor=COLOR_FILA_PAR)
        thin_border = Border(bottom=Side(style="thin", color=COLOR_BORDE))
        align_center = Alignment(horizontal="center", vertical="center")
        align_left   = Alignment(horizontal="left",   vertical="center")

        COLUMNAS    = 7
        fila_cursor = 1

        ws.merge_cells(f"A{fila_cursor}:{get_column_letter(COLUMNAS)}{fila_cursor}")
        celda = ws.cell(fila_cursor, 1, "Clínica Lichi — Listado de Pacientes")
        celda.font = font_titulo; celda.alignment = align_left
        ws.row_dimensions[fila_cursor].height = 20
        fila_cursor += 1

        ws.merge_cells(f"A{fila_cursor}:{get_column_letter(COLUMNAS)}{fila_cursor}")
        celda = ws.cell(fila_cursor, 1, f"Generado el {hoy.strftime('%d/%m/%Y')}  —  {len(filas)} registro{'s' if len(filas) != 1 else ''}")
        celda.font = font_meta; celda.alignment = align_left
        fila_cursor += 1

        if filtros_aplicados:
            fila_cursor += 1
            ws.merge_cells(f"A{fila_cursor}:{get_column_letter(COLUMNAS)}{fila_cursor}")
            ws.cell(fila_cursor, 1, "Filtros aplicados:").font = Font(bold=True, size=9, color="374151")
            fila_cursor += 1
            for etiqueta, valor in filtros_aplicados:
                ws.merge_cells(f"A{fila_cursor}:{get_column_letter(COLUMNAS)}{fila_cursor}")
                c = ws.cell(fila_cursor, 1, f"  {etiqueta}: {valor}")
                c.font = font_filtro; c.alignment = align_left
                fila_cursor += 1

        fila_cursor += 1
        for col, texto in enumerate(["N°","Nombre completo","Documento","Edad","Sexo","Teléfono","Responsable"], start=1):
            c = ws.cell(fila_cursor, col, texto)
            c.fill = fill_header; c.font = font_header
            c.alignment = align_center if col in (1, 4, 5) else align_left
        ws.row_dimensions[fila_cursor].height = 18
        fila_cursor += 1

        for fila in filas:
            es_par  = (fila["nro"] % 2 == 0)
            valores = [fila["nro"], fila["nombre"], fila["documento"],
                       fila["edad"], fila["sexo"], fila["telefono"], fila["responsable"]]
            for col, val in enumerate(valores, start=1):
                c = ws.cell(fila_cursor, col, val)
                if es_par: c.fill = fill_par
                c.border    = thin_border
                c.alignment = align_center if col in (1, 4, 5) else align_left
            ws.row_dimensions[fila_cursor].height = 15
            fila_cursor += 1

        for i, ancho in enumerate([6, 32, 16, 8, 12, 16, 32], start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="listado_pacientes_{hoy.strftime("%Y%m%d")}.xlsx"'
        return response

    # ── Dashboard mensual ─────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="dashboard-mensual",
            permission_classes=[IsAuthenticated])
    def dashboard_mensual(self, request):
        hoy       = timezone.localdate()
        ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]

        qs = Paciente.objects.filter(
            is_deleted=False,
            fecha_creacion__year=hoy.year,
            fecha_creacion__month=hoy.month,
        ).select_related("persona__pais", "persona__departamento")

        total_mes = qs.count()

        # Por día — incluye todos los días del mes, 0 para los sin registro
        por_dia_raw = {
            r["dia"]: r["total"]
            for r in qs.annotate(dia=TruncDate("fecha_creacion"))
                       .values("dia")
                       .annotate(total=Count("id"))
        }
        por_dia = [
            {
                "dia":       d,
                "fecha":     hoy.replace(day=d).strftime("%Y-%m-%d"),
                "total":     por_dia_raw.get(hoy.replace(day=d), 0),
                "es_futuro": hoy.replace(day=d) > hoy,
            }
            for d in range(1, ultimo_dia + 1)
        ]

        # Por semana (bloques de 7 días desde el día 1)
        por_semana = []
        sem = 1
        for inicio in range(0, ultimo_dia, 7):
            dias = list(range(inicio + 1, min(inicio + 8, ultimo_dia + 1)))
            total_sem = sum(
                por_dia_raw.get(hoy.replace(day=d), 0)
                for d in dias
                if hoy.replace(day=d) <= hoy
            )
            por_semana.append({
                "semana": sem,
                "label":  f"Sem {sem} ({dias[0]}–{dias[-1]})",
                "total":  total_sem,
            })
            sem += 1

        # Por sexo
        por_sexo = [
            {"sexo": r["sexo"], "label": ETIQUETAS_SEXO.get(r["sexo"], r["sexo"]), "total": r["total"]}
            for r in qs.values("sexo").annotate(total=Count("id")).order_by("-total")
        ]

        # Por grupo etario
        GRUPOS = [("0–12", 0, 12), ("13–17", 13, 17), ("18–29", 18, 29),
                  ("30–44", 30, 44), ("45–59", 45, 59), ("60+", 60, 999)]
        grupos_cnt = {lbl: 0 for lbl, _, _ in GRUPOS}
        sin_fecha  = 0
        qs_fechas = Paciente.objects.filter(
            is_deleted=False,
            fecha_creacion__year=hoy.year,
            fecha_creacion__month=hoy.month,
        ).values_list("fecha_nacimiento", flat=True)
        for fecha_nac in qs_fechas:
            if not fecha_nac:
                sin_fecha += 1
                continue
            edad = hoy.year - fecha_nac.year - (
                (hoy.month, hoy.day) < (fecha_nac.month, fecha_nac.day)
            )
            for lbl, lo, hi in GRUPOS:
                if lo <= edad <= hi:
                    grupos_cnt[lbl] += 1
                    break
        por_grupo_etario = [{"label": f"{lbl} años", "total": grupos_cnt[lbl]} for lbl, _, _ in GRUPOS]
        if sin_fecha:
            por_grupo_etario.append({"label": "Sin fecha", "total": sin_fecha})

        # Por departamento — top 5 + otros, indicando el país
        depto_qs = (
            qs.values("persona__departamento_id",
                      "persona__departamento__descripcion",
                      "persona__pais__descripcion")
              .annotate(total=Count("id"))
              .order_by("-total")
        )
        todos         = list(depto_qs)
        otros_total   = sum(r["total"] for r in todos[5:])
        por_depto = [
            {
                "id":    r["persona__departamento_id"],
                "label": r["persona__departamento__descripcion"] or "Sin departamento",
                "pais":  r["persona__pais__descripcion"] or "",
                "total": r["total"],
            }
            for r in todos[:5]
        ]
        if otros_total:
            por_depto.append({"id": None, "label": "Otros", "pais": "", "total": otros_total})

        # Tendencia — últimos 6 meses (incluye el mes actual)
        tendencia = []
        for i in range(5, -1, -1):
            mes  = hoy.month - i
            anio = hoy.year
            while mes <= 0:
                mes  += 12
                anio -= 1
            cnt = Paciente.objects.filter(
                is_deleted=False,
                fecha_creacion__year=anio,
                fecha_creacion__month=mes,
            ).count()
            tendencia.append({
                "periodo": f"{anio}-{mes:02d}",
                "label":   MESES_CORTO[mes - 1],
                "total":   cnt,
            })

        return Response({
            "mes_label":        f"{MESES[hoy.month - 1]} {hoy.year}",
            "total_mes":        total_mes,
            "hoy":              hoy.strftime("%Y-%m-%d"),
            "ultimo_dia":       ultimo_dia,
            "por_dia":          por_dia,
            "por_semana":       por_semana,
            "por_sexo":         por_sexo,
            "por_grupo_etario": por_grupo_etario,
            "por_departamento": por_depto,
            "tendencia_6meses": tendencia,
        })
