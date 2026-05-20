from rest_framework import viewsets, filters
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from datetime import date, timedelta
from .models import Consulta
from .serializers import ConsultaSerializer, ConsultaListSerializer
from config.pagination import StandardPagination
from apps.administracion.auditoria.mixins import AuditoriaMixin
from apps.core.permissions import IsAdminRole


class ConsultaViewSet(AuditoriaMixin, viewsets.ModelViewSet):
    pagination_class = StandardPagination
    filter_backends  = [filters.OrderingFilter]
    ordering_fields  = ['agenda__fecha', 'hora_desde']
    ordering         = ['-agenda__fecha', 'hora_desde']

    def get_permissions(self):
        if self.action in ('destroy', 'eliminados'):
            return [IsAuthenticated(), IsAdminRole()]
        return [IsAuthenticated()]

    def _consultas_agrupadas(self, request):
        from collections import OrderedDict
        qs = (
            Consulta.objects
            .filter(is_deleted=False, agenda__is_deleted=False)
            .select_related(
                'agenda__horario_prestador__persona_rrhh__persona',
                'agenda__paciente__persona',
                'evento_clinico',
            )
            .prefetch_related('agenda__horario_prestador__especialidades')
            .order_by('agenda__fecha', 'agenda__hora_desde')
        )
        params = request.query_params
        persona_rrhh   = params.get('persona_rrhh')
        especialidad   = params.get('especialidad')
        evento_clinico = params.get('evento_clinico')
        paciente       = params.get('paciente')
        fecha_desde    = params.get('fecha_desde')
        fecha_hasta    = params.get('fecha_hasta')

        if persona_rrhh:   qs = qs.filter(agenda__horario_prestador__persona_rrhh_id=persona_rrhh)
        if especialidad:   qs = qs.filter(agenda__horario_prestador__especialidades__id=especialidad).distinct()
        if evento_clinico: qs = qs.filter(evento_clinico_id=evento_clinico)
        if paciente:       qs = qs.filter(agenda__paciente_id=paciente)
        if fecha_desde:    qs = qs.filter(agenda__fecha__gte=fecha_desde)
        if fecha_hasta:    qs = qs.filter(agenda__fecha__lte=fecha_hasta)

        agrupado = OrderedDict()
        for c in qs:
            esps   = list(c.agenda.horario_prestador.especialidades.all())
            clave  = esps[0].descripcion if esps else 'Sin especialidad'
            if clave not in agrupado:
                agrupado[clave] = []
            agrupado[clave].append({
                'fecha':          str(c.agenda.fecha),
                'paciente':       c.agenda.paciente.persona.razon_social if c.agenda.paciente else '—',
                'prestador':      c.agenda.horario_prestador.persona_rrhh.persona.razon_social,
                'evento_clinico': c.evento_clinico.tipo_evento if c.evento_clinico else '—',
                'diagnostico':    (c.diagnostico or '').strip()[:100] or '—',
                'estado':         c.get_estado_display(),
            })
        return sorted(
            [{'especialidad': k, 'consultas': v} for k, v in agrupado.items()],
            key=lambda x: x['especialidad'],
        )

    def get_queryset(self):
        qs = Consulta.objects.filter(is_deleted=False).select_related(
            'agenda__horario_prestador__persona_rrhh__persona',
            'agenda__paciente__persona',
            'agenda__paciente__responsable__persona',
            'evento_clinico',
        ).prefetch_related(
            'agenda__horario_prestador__especialidades',
        )

        rol               = self.request.auth.get('rol')                   if self.request.auth else None
        persona_rrhh_id   = self.request.auth.get('persona_rrhh_id')       if self.request.auth else None
        medicos_asignados = self.request.auth.get('medicos_asignados', []) if self.request.auth else []

        if rol == 'medico':
            if not persona_rrhh_id:
                return qs.none()
            qs = qs.filter(agenda__horario_prestador__persona_rrhh_id=persona_rrhh_id)
        elif rol == 'secretaria_medico':
            if not medicos_asignados:
                return qs.none()
            qs = qs.filter(agenda__horario_prestador__persona_rrhh_id__in=medicos_asignados)
        else:
            persona_rrhh = self.request.query_params.get('persona_rrhh')
            if persona_rrhh:
                qs = qs.filter(agenda__horario_prestador__persona_rrhh_id=persona_rrhh)

        params   = self.request.query_params
        fecha    = params.get('fecha')
        estado   = params.get('estado')
        paciente = params.get('paciente')

        if fecha:    qs = qs.filter(agenda__fecha=fecha)
        if estado:   qs = qs.filter(estado=estado)
        if paciente: qs = qs.filter(agenda__paciente_id=paciente)

        return qs

    def get_serializer_class(self):
        if self.action in ['list', 'retrieve', 'eliminados']:
            return ConsultaListSerializer
        return ConsultaSerializer

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.estado in (Consulta.Estado.FINALIZADA, Consulta.Estado.ANULADA):
            rol = request.auth.get('rol') if request.auth else None
            if rol not in ('medico', 'admin'):
                raise ValidationError(
                    {'estado': 'Solo médicos y administradores pueden modificar consultas finalizadas o anuladas.'}
                )
            if timezone.now() > instance.fecha_modificacion + timedelta(hours=24):
                raise ValidationError(
                    {'estado': 'Esta consulta no puede modificarse: han pasado más de 24 horas desde su cierre.'}
                )
        return super().partial_update(request, *args, **kwargs)

    def perform_destroy(self, instance):
        from apps.clinica.configuracion.documentos.models import DocumentoDigPaciente
        if DocumentoDigPaciente.objects.filter(consulta=instance, is_deleted=False).exists():
            raise ValidationError('No se puede eliminar: la consulta tiene documentos asociados.')
        super().perform_destroy(instance)

    @action(detail=False, methods=['get'], url_path='eliminados')
    def eliminados(self, request):
        qs = Consulta.objects.filter(is_deleted=True).select_related(
            'agenda__horario_prestador__persona_rrhh__persona',
            'agenda__paciente__persona',
            'agenda__paciente__responsable__persona',
            'evento_clinico',
        ).order_by('-fecha_eliminacion')
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        return Response(self.get_serializer(qs, many=True).data)

    @action(detail=True, methods=['post'], url_path='iniciar')
    def iniciar(self, request, pk=None):
        consulta = self.get_object()

        if consulta.estado != Consulta.Estado.EN_ESPERA:
            raise ValidationError(
                {'estado': 'Solo se puede iniciar una consulta en espera. Estado actual: ' + str(consulta.estado) + '.'}
            )

        consulta.estado             = Consulta.Estado.EN_CONSULTA
        consulta.hora_desde         = timezone.localtime().time()
        consulta.id_usu_modificator = request.user
        consulta.save()

        return Response(ConsultaListSerializer(consulta).data)

    @action(detail=True, methods=['post'], url_path='anular')
    def anular(self, request, pk=None):
        consulta = self.get_object()

        if consulta.estado != Consulta.Estado.EN_CONSULTA:
            raise ValidationError(
                {'estado': 'Solo se puede anular una consulta en curso. Estado actual: ' + str(consulta.estado) + '.'}
            )

        consulta.estado             = Consulta.Estado.ANULADA
        consulta.hora_hasta         = timezone.localtime().time()
        consulta.id_usu_modificator = request.user
        consulta.save()

        return Response(ConsultaListSerializer(consulta).data)

    @action(detail=True, methods=['post'], url_path='finalizar')
    def finalizar(self, request, pk=None):
        consulta = self.get_object()

        if consulta.estado != Consulta.Estado.EN_CONSULTA:
            raise ValidationError(
                {'estado': 'Solo se puede finalizar una consulta en curso. Estado actual: ' + str(consulta.estado) + '.'}
            )

        consulta.estado             = Consulta.Estado.FINALIZADA
        consulta.hora_hasta         = timezone.localtime().time()
        consulta.id_usu_modificator = request.user
        consulta.save()

        from apps.clinica.agenda.models import Agenda
        agenda = consulta.agenda
        agenda.estado             = Agenda.Estado.REALIZADO
        agenda.id_usu_modificator = request.user
        agenda.save()

        return Response(ConsultaListSerializer(consulta).data)

    @action(detail=False, methods=['get'], url_path='stats-hoy')
    def stats_hoy(self, request):
        hoy = timezone.localtime().date()
        qs  = Consulta.objects.filter(agenda__fecha=hoy, is_deleted=False)
        return Response({
            'total':       qs.count(),
            'en_espera':   qs.filter(estado=Consulta.Estado.EN_ESPERA).count(),
            'en_consulta': qs.filter(estado=Consulta.Estado.EN_CONSULTA).count(),
            'finalizadas': qs.filter(estado=Consulta.Estado.FINALIZADA).count(),
        })

    @action(detail=False, methods=['get'], url_path='reporte-consultas')
    def reporte_consultas(self, request):
        grupos = self._consultas_agrupadas(request)
        total  = sum(len(g['consultas']) for g in grupos)
        hoy    = date.today()
        ctx    = {'grupos': grupos, 'fecha': hoy, 'total': total}
        html   = render_to_string('informes/consultas_lista.html', ctx, request=request)
        try:
            import weasyprint
            pdf = weasyprint.HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
        except Exception as e:
            return HttpResponse(f'Error generando PDF: {e}', status=500)
        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = 'inline; filename="consultas.pdf"'
        return resp

    @action(detail=False, methods=['get'], url_path='reporte-consultas-excel')
    def reporte_consultas_excel(self, request):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse('openpyxl no está instalado.', status=500)

        grupos = self._consultas_agrupadas(request)
        total  = sum(len(g['consultas']) for g in grupos)
        hoy    = date.today()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Listado de Consultas'

        fill_header  = PatternFill('solid', fgColor='1A3A5C')
        fill_grupo   = PatternFill('solid', fgColor='EEF2F7')
        font_header  = Font(color='FFFFFF', bold=True, size=10)
        font_grupo   = Font(color='1A3A5C', bold=True, size=10)
        font_titulo  = Font(color='1A3A5C', bold=True, size=13)
        font_meta    = Font(color='555555', size=9)
        fill_par     = PatternFill('solid', fgColor='F8FAFC')
        thin_border  = Border(bottom=Side(style='thin', color='E8EDF2'))
        align_center = Alignment(horizontal='center', vertical='center')
        align_left   = Alignment(horizontal='left',   vertical='center')

        COLS = 6
        fila = 1

        ws.merge_cells(f'A{fila}:{get_column_letter(COLS)}{fila}')
        c = ws.cell(fila, 1, 'Clínica Lichi — Listado de Consultas')
        c.font = font_titulo; c.alignment = align_left
        ws.row_dimensions[fila].height = 20
        fila += 1

        ws.merge_cells(f'A{fila}:{get_column_letter(COLS)}{fila}')
        c = ws.cell(fila, 1, f"Generado el {hoy.strftime('%d/%m/%Y')}  —  {total} registro{'s' if total != 1 else ''}")
        c.font = font_meta; c.alignment = align_left
        fila += 2

        headers = ['N°', 'Fecha', 'Paciente', 'Prestador', 'Evento Clínico', 'Estado']
        for col, txt in enumerate(headers, 1):
            c = ws.cell(fila, col, txt)
            c.fill = fill_header; c.font = font_header
            c.alignment = align_center if col in (1, 2, 6) else align_left
        ws.row_dimensions[fila].height = 18
        fila += 1

        for grupo in grupos:
            ws.merge_cells(f'A{fila}:{get_column_letter(COLS)}{fila}')
            c = ws.cell(fila, 1, grupo['especialidad'])
            c.fill = fill_grupo; c.font = font_grupo; c.alignment = align_left
            ws.row_dimensions[fila].height = 16
            fila += 1

            for nro, con in enumerate(grupo['consultas'], 1):
                es_par = (nro % 2 == 0)
                vals   = [nro, con['fecha'], con['paciente'], con['prestador'], con['evento_clinico'], con['estado']]
                for col, val in enumerate(vals, 1):
                    c = ws.cell(fila, col, val)
                    if es_par: c.fill = fill_par
                    c.border    = thin_border
                    c.alignment = align_center if col in (1, 2, 6) else align_left
                ws.row_dimensions[fila].height = 15
                fila += 1

        for i, ancho in enumerate([4, 11, 24, 24, 20, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="consultas_{hoy.strftime("%Y%m%d")}.xlsx"'
        return resp

    @action(detail=False, methods=['get'], url_path='dashboard-consultas')
    def dashboard_consultas(self, request):
        import calendar
        from django.db.models import Count, F

        hoy        = date.today()
        inicio_mes = hoy.replace(day=1)

        qs_base = Consulta.objects.filter(is_deleted=False, agenda__is_deleted=False)
        qs_mes  = qs_base.filter(agenda__fecha__gte=inicio_mes, agenda__fecha__lte=hoy)

        total_mes  = qs_mes.count()
        por_estado = {
            'en_espera':   qs_mes.filter(estado='en_espera').count(),
            'en_consulta': qs_mes.filter(estado='en_consulta').count(),
            'finalizada':  qs_mes.filter(estado='finalizada').count(),
            'anulada':     qs_mes.filter(estado='anulada').count(),
        }

        top_prestadores = list(
            qs_mes
            .values(nombre=F('agenda__horario_prestador__persona_rrhh__persona__razon_social'))
            .annotate(total=Count('id'))
            .order_by('-total')[:5]
        )

        por_especialidad = list(
            qs_mes
            .filter(agenda__horario_prestador__especialidades__isnull=False)
            .values(especialidad=F('agenda__horario_prestador__especialidades__descripcion'))
            .annotate(total=Count('id', distinct=True))
            .order_by('-total')[:8]
        )

        MESES_ES = ['', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        comparativa = []
        for i in range(5, -1, -1):
            mes_ref  = hoy.month - i
            anio_ref = hoy.year
            while mes_ref <= 0:
                mes_ref  += 12
                anio_ref -= 1
            inicio = date(anio_ref, mes_ref, 1)
            fin    = date(anio_ref, mes_ref, calendar.monthrange(anio_ref, mes_ref)[1])
            if fin > hoy:
                fin = hoy
            comparativa.append({
                'mes':   mes_ref,
                'anio':  anio_ref,
                'label': f"{MESES_ES[mes_ref]} {str(anio_ref)[2:]}",
                'total': qs_base.filter(agenda__fecha__gte=inicio, agenda__fecha__lte=fin).count(),
            })

        return Response({
            'total_mes':           total_mes,
            'por_estado':          por_estado,
            'top_prestadores':     top_prestadores,
            'por_especialidad':    por_especialidad,
            'comparativa_6_meses': comparativa,
        })

    @action(detail=False, methods=['get'], url_path='historia-clinica')
    def historia_clinica(self, request):
        import weasyprint
        from rest_framework import status as drf_status
        from apps.clinica.paciente.models import Paciente
        from apps.clinica.configuracion.documentos.models import DocumentoDigPaciente

        paciente_id = request.query_params.get('paciente', '')
        if not paciente_id:
            return Response({'error': 'Debe especificar un paciente.'}, status=drf_status.HTTP_400_BAD_REQUEST)

        try:
            paciente = Paciente.objects.select_related(
                'persona__tipo_documento',
                'persona__pais',
                'persona__departamento',
                'persona__ciudad',
                'responsable__persona',
            ).get(id=paciente_id, is_deleted=False)
        except Paciente.DoesNotExist:
            return Response({'error': 'Paciente no encontrado.'}, status=drf_status.HTTP_404_NOT_FOUND)

        consultas = (
            Consulta.objects
            .filter(agenda__paciente=paciente, is_deleted=False, agenda__is_deleted=False)
            .select_related(
                'agenda__horario_prestador__persona_rrhh__persona',
                'agenda__horario_prestador__consultorio',
                'evento_clinico',
            )
            .prefetch_related('agenda__horario_prestador__especialidades')
            .order_by('-agenda__fecha', '-agenda__hora_desde')
        )

        # Pre-cargar documentos de todas las consultas
        ids_consulta = [c.id for c in consultas]
        docs_qs = DocumentoDigPaciente.objects.filter(
            consulta_id__in=ids_consulta, is_deleted=False
        ).select_related('tipo_doc_dig').values('consulta_id', 'tipo_doc_dig__descripcion', 'filename')
        docs_map = {}
        for d in docs_qs:
            cid = d['consulta_id']
            if cid not in docs_map:
                docs_map[cid] = []
            docs_map[cid].append(d['filename'] or d['tipo_doc_dig__descripcion'] or '—')

        consultas_data = []
        for c in consultas:
            hp  = c.agenda.horario_prestador
            esps = list(hp.especialidades.all()) if hp else []
            consultas_data.append({
                'fecha':          c.agenda.fecha.strftime('%d/%m/%Y') if c.agenda.fecha else '—',
                'hora':           c.agenda.hora_desde.strftime('%H:%M') if c.agenda.hora_desde else '—',
                'medico':         hp.persona_rrhh.persona.razon_social if hp and hp.persona_rrhh and hp.persona_rrhh.persona else '—',
                'especialidad':   ', '.join(e.descripcion for e in esps) or '—',
                'consultorio':    hp.consultorio.descripcion if hp and hp.consultorio else '—',
                'evento':         c.evento_clinico.tipo_evento if c.evento_clinico else '—',
                'motivo':         (c.motivo_consulta or '').strip(),
                'diagnostico':    (c.diagnostico or '').strip(),
                'tratamiento':    (c.tratamiento or '').strip(),
                'indicaciones':   (c.indicaciones or '').strip(),
                'proxima_cita':   c.proxima_cita.strftime('%d/%m/%Y') if c.proxima_cita else '',
                'estado':         c.get_estado_display(),
                'documentos':     docs_map.get(c.id, []),
            })

        hoy = timezone.localdate()
        edad = None
        if paciente.persona and paciente.persona.fecha_nacimiento:
            fn = paciente.persona.fecha_nacimiento
            edad = hoy.year - fn.year - ((hoy.month, hoy.day) < (fn.month, fn.day))

        html = render_to_string('informes/historia_clinica.html', {
            'paciente':       paciente,
            'edad':           edad,
            'consultas':      consultas_data,
            'total':          len(consultas_data),
            'fecha_emision':  hoy.strftime('%d/%m/%Y'),
        })
        pdf = weasyprint.HTML(string=html).write_pdf()
        response = HttpResponse(pdf, content_type='application/pdf')
        safe_name = (paciente.persona.razon_social or 'paciente').replace(' ', '_').lower()[:25]
        response['Content-Disposition'] = f'inline; filename="historia_{safe_name}.pdf"'
        return response
