from datetime import date, datetime, timedelta
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from rest_framework import viewsets, filters
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from apps.administracion.auditoria.mixins import AuditoriaMixin
from apps.core.permissions import IsAdminRole, IsAdminOrRecepcionista, IsAdminOrRecepcionistaOrSecretaria
from .models import HorarioPrestador
from .serializers import HorarioPrestadorSerializer, HorarioPrestadorListSerializer
from config.pagination import StandardPagination


class HorarioPrestadorViewSet(AuditoriaMixin, viewsets.ModelViewSet):
    pagination_class = StandardPagination
    filter_backends  = [filters.OrderingFilter]
    ordering_fields  = ['dia_semana__id', 'hora_desde', 'fecha_creacion']

    def get_permissions(self):
        if self.action in ('list', 'retrieve', 'reporte_horarios', 'reporte_horarios_excel'):
            return [IsAuthenticated()]
        rol = self.request.auth.get('rol') if self.request.auth else None
        if rol in ('medico', 'secretaria_medico'):
            if self.action == 'eliminados':
                return [IsAuthenticated(), IsAdminRole()]
            return [IsAuthenticated()]
        if self.action in ('destroy', 'eliminados'):
            return [IsAuthenticated(), IsAdminRole()]
        if self.action == 'generar':
            return [IsAuthenticated(), IsAdminOrRecepcionistaOrSecretaria()]
        return [IsAuthenticated(), IsAdminOrRecepcionista()]

    def _check_medico_ownership(self, persona_rrhh):
        if not self.request.auth:
            return
        rol = self.request.auth.get('rol')
        obj_id = persona_rrhh.id if hasattr(persona_rrhh, 'id') else int(persona_rrhh)
        if rol == 'medico':
            jwt_id = self.request.auth.get('persona_rrhh_id')
            if not jwt_id:
                raise ValidationError({'persona_rrhh': 'No se pudo verificar el prestador del token.'})
            if obj_id != int(jwt_id):
                raise ValidationError({'persona_rrhh': 'Solo podés gestionar tus propios horarios.'})
        elif rol == 'secretaria_medico':
            asignados = self.request.auth.get('medicos_asignados', [])
            if not asignados or obj_id not in [int(x) for x in asignados]:
                raise ValidationError({'persona_rrhh': 'Solo podés gestionar los horarios de tu médico asignado.'})

    def perform_create(self, serializer):
        self._check_medico_ownership(serializer.validated_data.get('persona_rrhh'))
        super().perform_create(serializer)

    def perform_update(self, serializer):
        self._check_medico_ownership(serializer.instance.persona_rrhh)
        super().perform_update(serializer)

    def get_queryset(self):
        qs = HorarioPrestador.objects.filter(
            is_deleted=False,
            persona_rrhh__is_deleted=False,
        ).select_related(
            'persona_rrhh__persona',
            'persona_rrhh__persona__tipo_documento',
            'persona_rrhh__persona__pais',
            'persona_rrhh__persona__departamento',
            'persona_rrhh__persona__ciudad',
            'dia_semana',
        ).prefetch_related(
            'especialidades',
        )

        rol               = self.request.auth.get('rol')                   if self.request.auth else None
        persona_rrhh_id   = self.request.auth.get('persona_rrhh_id')       if self.request.auth else None
        medicos_asignados = self.request.auth.get('medicos_asignados', []) if self.request.auth else []

        if rol == 'medico':
            if not persona_rrhh_id:
                return qs.none()
            return qs.filter(persona_rrhh_id=persona_rrhh_id)

        if rol == 'secretaria_medico':
            if not medicos_asignados:
                return qs.none()
            return qs.filter(persona_rrhh_id__in=medicos_asignados)

        persona_rrhh = self.request.query_params.get('persona_rrhh')
        estado       = self.request.query_params.get('estado')
        if persona_rrhh:
            qs = qs.filter(persona_rrhh_id=persona_rrhh)
        if estado:
            qs = qs.filter(estado=estado)
        return qs

    def get_serializer_class(self):
        if self.action in ['list', 'retrieve', 'eliminados']:
            return HorarioPrestadorListSerializer
        return HorarioPrestadorSerializer

    def perform_destroy(self, instance):
        self._check_medico_ownership(instance.persona_rrhh)
        from apps.clinica.agenda.models import Agenda
        tiene_turnos = Agenda.objects.filter(
            horario_prestador=instance,
            is_deleted=False,
            estado__in=[
                Agenda.Estado.DISPONIBLE,
                Agenda.Estado.OCUPADO,
                Agenda.Estado.REALIZADO,
            ],
        ).exists()
        if tiene_turnos:
            raise ValidationError(
                'No se puede eliminar un horario con turnos activos (disponible, ocupado o realizado). '
                'Primero cancelá o finalizá todos sus turnos.'
            )
        instance.especialidades.clear()
        super().perform_destroy(instance)

    @action(detail=False, methods=['get'], url_path='eliminados')
    def eliminados(self, request):
        qs = HorarioPrestador.objects.filter(
            is_deleted=True
        ).select_related(
            'persona_rrhh__persona',
            'persona_rrhh__persona__tipo_documento',
            'persona_rrhh__persona__pais',
            'persona_rrhh__persona__departamento',
            'persona_rrhh__persona__ciudad',
            'dia_semana',
        ).prefetch_related(
            'especialidades',
        )
        serializer = HorarioPrestadorListSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='generar')
    def generar(self, request, pk=None):
        from apps.clinica.agenda.models import Agenda

        horario       = self.get_object()
        self._check_medico_ownership(horario.persona_rrhh)
        fecha_desde_s = request.data.get('fecha_desde')
        fecha_hasta_s = request.data.get('fecha_hasta')

        if not fecha_desde_s or not fecha_hasta_s:
            return Response(
                {'error': 'fecha_desde y fecha_hasta son requeridos.'},
                status=400,
            )
        try:
            fecha_desde = date.fromisoformat(fecha_desde_s)
            fecha_hasta = date.fromisoformat(fecha_hasta_s)
        except ValueError:
            return Response(
                {'error': 'Formato de fecha inválido. Use YYYY-MM-DD.'},
                status=400,
            )

        now_local  = timezone.localtime()
        hoy        = now_local.date()
        hora_ahora = now_local.time().replace(microsecond=0)

        # Ajustar fecha_desde al día de hoy si viene en el pasado
        if fecha_desde < hoy:
            fecha_desde = hoy

        if fecha_hasta < fecha_desde:
            return Response(
                {'error': 'La fecha "Hasta" no puede ser menor a la fecha actual.'},
                status=400,
            )

        creados   = 0
        omitidos  = 0
        detalle   = []
        intervalo = timedelta(minutes=horario.intervalo)

        for i in range((fecha_hasta - fecha_desde).days + 1):
            fecha_actual = fecha_desde + timedelta(days=i)

            if horario.excepcion:
                if horario.fecha_excepcion != fecha_actual:
                    continue
            else:
                if horario.dia_semana.id != fecha_actual.weekday() + 1:
                    continue

            turnos_creados  = 0
            turnos_omitidos = 0
            hora_actual = datetime.combine(fecha_actual, horario.hora_desde)
            hora_limite = datetime.combine(fecha_actual, horario.hora_hasta)

            # Para el día de hoy, avanzar al primer slot que no haya pasado
            if fecha_actual == hoy:
                ahora_dt = datetime.combine(hoy, hora_ahora)
                while hora_actual <= ahora_dt and hora_actual < hora_limite:
                    hora_actual += intervalo

            while hora_actual < hora_limite:
                hora_fin = hora_actual + intervalo

                existe = Agenda.objects.filter(
                    horario_prestador=horario,
                    fecha=fecha_actual,
                    hora_desde=hora_actual.time(),
                    is_deleted=False,
                ).exists()

                if not existe:
                    Agenda.objects.create(
                        horario_prestador = horario,
                        fecha             = fecha_actual,
                        hora_desde        = hora_actual.time(),
                        hora_hasta        = hora_fin.time(),
                        estado            = Agenda.Estado.DISPONIBLE,
                        id_usu_creator    = request.user,
                    )
                    turnos_creados += 1
                    creados        += 1
                else:
                    turnos_omitidos += 1
                    omitidos        += 1

                hora_actual += intervalo

            detalle.append({
                'fecha':           str(fecha_actual),
                'dia':             horario.dia_semana.descripcion,
                'turnos_creados':  turnos_creados,
                'turnos_omitidos': turnos_omitidos,
            })

        return Response({
            'creados':  creados,
            'omitidos': omitidos,
            'detalle':  detalle,
        })

    def _horarios_agrupados(self, request):
        qs = (
            HorarioPrestador.objects
            .filter(is_deleted=False, persona_rrhh__is_deleted=False)
            .select_related('persona_rrhh__persona', 'dia_semana')
            .prefetch_related('especialidades')
            .order_by('persona_rrhh__persona__razon_social', 'dia_semana__id', 'hora_desde')
        )
        persona_rrhh_id = request.query_params.get('persona_rrhh')
        dia_semana_id   = request.query_params.get('dia_semana')
        if persona_rrhh_id:
            qs = qs.filter(persona_rrhh_id=persona_rrhh_id)
        if dia_semana_id:
            qs = qs.filter(dia_semana_id=dia_semana_id)

        from collections import OrderedDict
        agrupado = OrderedDict()
        for h in qs:
            nombre = h.persona_rrhh.persona.razon_social
            dia = h.dia_semana.descripcion if h.dia_semana else '—'
            if h.excepcion and h.fecha_excepcion:
                dia = f'Excepción ({h.fecha_excepcion})'
            if nombre not in agrupado:
                agrupado[nombre] = []
            agrupado[nombre].append({
                'dia':            dia,
                'desde':          str(h.hora_desde)[:5],
                'hasta':          str(h.hora_hasta)[:5],
                'intervalo':      f'{h.intervalo} min',
                'especialidades': ', '.join(e.descripcion for e in h.especialidades.all()) or '—',
                'estado':         h.get_estado_display(),
            })
        return [{'nombre': nombre, 'horarios': hs} for nombre, hs in agrupado.items()]

    @action(detail=False, methods=['get'], url_path='reporte-horarios')
    def reporte_horarios(self, request):
        prestadores = self._horarios_agrupados(request)
        total       = sum(len(p['horarios']) for p in prestadores)
        hoy         = date.today()
        contexto    = {'prestadores': prestadores, 'fecha': hoy, 'total': total}
        html        = render_to_string('informes/horario_prestador_lista.html', contexto, request=request)
        try:
            import weasyprint
            pdf_bytes = weasyprint.HTML(
                string=html, base_url=request.build_absolute_uri('/')
            ).write_pdf()
        except Exception as e:
            return HttpResponse(f'Error generando PDF: {e}', status=500)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="horarios_prestadores.pdf"'
        return response

    @action(detail=False, methods=['get'], url_path='reporte-horarios-excel')
    def reporte_horarios_excel(self, request):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse('openpyxl no está instalado.', status=500)

        prestadores = self._horarios_agrupados(request)
        total       = sum(len(p['horarios']) for p in prestadores)
        hoy         = date.today()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Horarios de Prestadores'

        fill_header    = PatternFill('solid', fgColor='1A3A5C')
        fill_prestador = PatternFill('solid', fgColor='EEF2F7')
        font_header    = Font(color='FFFFFF', bold=True, size=10)
        font_prestador = Font(color='1A3A5C', bold=True, size=10)
        font_titulo    = Font(color='1A3A5C', bold=True, size=13)
        font_meta      = Font(color='555555', size=9)
        fill_par       = PatternFill('solid', fgColor='F8FAFC')
        thin_border    = Border(bottom=Side(style='thin', color='E8EDF2'))
        align_center   = Alignment(horizontal='center', vertical='center')
        align_left     = Alignment(horizontal='left',   vertical='center')

        COLUMNAS    = 7
        fila_cursor = 1

        ws.merge_cells(f'A{fila_cursor}:{get_column_letter(COLUMNAS)}{fila_cursor}')
        c = ws.cell(fila_cursor, 1, 'Clínica Lichi — Horarios de Prestadores')
        c.font = font_titulo; c.alignment = align_left
        ws.row_dimensions[fila_cursor].height = 20
        fila_cursor += 1

        ws.merge_cells(f'A{fila_cursor}:{get_column_letter(COLUMNAS)}{fila_cursor}')
        c = ws.cell(fila_cursor, 1, f"Generado el {hoy.strftime('%d/%m/%Y')}  —  {total} registro{'s' if total != 1 else ''}")
        c.font = font_meta; c.alignment = align_left
        fila_cursor += 2

        headers = ['N°', 'Día / Fecha', 'Desde', 'Hasta', 'Intervalo', 'Especialidades', 'Estado']
        for col, texto in enumerate(headers, start=1):
            c = ws.cell(fila_cursor, col, texto)
            c.fill = fill_header; c.font = font_header
            c.alignment = align_center if col in (1, 3, 4, 5) else align_left
        ws.row_dimensions[fila_cursor].height = 18
        fila_cursor += 1

        for prestador in prestadores:
            ws.merge_cells(f'A{fila_cursor}:{get_column_letter(COLUMNAS)}{fila_cursor}')
            c = ws.cell(fila_cursor, 1, prestador['nombre'])
            c.fill = fill_prestador; c.font = font_prestador; c.alignment = align_left
            ws.row_dimensions[fila_cursor].height = 16
            fila_cursor += 1

            for nro, h in enumerate(prestador['horarios'], start=1):
                es_par  = (nro % 2 == 0)
                valores = [nro, h['dia'], h['desde'], h['hasta'], h['intervalo'], h['especialidades'], h['estado']]
                for col, val in enumerate(valores, start=1):
                    c = ws.cell(fila_cursor, col, val)
                    if es_par: c.fill = fill_par
                    c.border    = thin_border
                    c.alignment = align_center if col in (1, 3, 4, 5) else align_left
                ws.row_dimensions[fila_cursor].height = 15
                fila_cursor += 1

        for i, ancho in enumerate([5, 20, 7, 7, 10, 34, 10], start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="horarios_prestadores_{hoy.strftime("%Y%m%d")}.xlsx"'
        return response
