from datetime import date
from django.template.loader import render_to_string
from django.http import HttpResponse
from rest_framework import viewsets, filters
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from apps.administracion.auditoria.mixins import AuditoriaMixin
from apps.core.permissions import IsAdminRole
from apps.administracion.persona.models import Persona
from apps.administracion.persona.serializers import PersonaListSerializer
from config.pagination import StandardPagination
from .models import PersonaRRHH
from .serializers import PersonaRRHHSerializer, PersonaRRHHListSerializer


class PersonaRRHHViewSet(AuditoriaMixin, viewsets.ModelViewSet):
    pagination_class = StandardPagination
    filter_backends  = [filters.SearchFilter, filters.OrderingFilter]
    search_fields    = ["persona__razon_social", "persona__nro_documento", "nro_matricula"]
    ordering_fields  = ["persona__razon_social", "cargo", "estado", "fecha_creacion"]

    def get_permissions(self):
        if self.action in ('list', 'retrieve', 'buscar', 'validar_matricula', 'reporte_lista', 'reporte_lista_excel'):
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsAdminRole()]

    def get_queryset(self):
        qs = PersonaRRHH.objects.filter(
            is_deleted=False
        ).select_related(
            "persona__tipo_documento",
            "persona__pais",
            "persona__departamento",
            "persona__ciudad",
        ).prefetch_related("especialidades")

        if not self.request.auth:
            return qs

        rol               = self.request.auth.get('rol')
        persona_rrhh_id   = self.request.auth.get('persona_rrhh_id')
        medicos_asignados = self.request.auth.get('medicos_asignados', [])

        if rol == 'medico':
            if not persona_rrhh_id:
                return qs.none()
            return qs.filter(id=persona_rrhh_id)

        if rol == 'secretaria_medico':
            if not medicos_asignados:
                return qs.none()
            return qs.filter(id__in=medicos_asignados)

        return qs

    def get_serializer_class(self):
        if self.action in ["list", "retrieve", "eliminados", "buscar"]:
            return PersonaRRHHListSerializer
        return PersonaRRHHSerializer

    def perform_destroy(self, instance):
        from django.utils import timezone as tz
        from apps.clinica.agenda.models import Agenda
        from apps.clinica.configuracion.horario_prestador.models import HorarioPrestador
        tiene_turnos = Agenda.objects.filter(
            horario_prestador__persona_rrhh=instance,
            is_deleted=False,
            estado__in=[
                Agenda.Estado.DISPONIBLE,
                Agenda.Estado.OCUPADO,
                Agenda.Estado.REALIZADO,
            ],
        ).exists()
        if tiene_turnos:
            raise ValidationError(
                "No se puede eliminar un prestador con turnos activos. "
                "Primero cancelá o finalizá todos sus turnos."
            )
        ahora = tz.now()
        uid   = self.request.user.pk
        Agenda.objects.filter(
            horario_prestador__persona_rrhh=instance,
            is_deleted=False,
        ).update(is_deleted=True, fecha_eliminacion=ahora, id_usu_modificator_id=uid)
        HorarioPrestador.objects.filter(
            persona_rrhh=instance,
            is_deleted=False,
        ).update(is_deleted=True, fecha_eliminacion=ahora, id_usu_modificator_id=uid)
        instance.especialidades.clear()
        super().perform_destroy(instance)

    @action(detail=False, methods=["get"], url_path="eliminados")
    def eliminados(self, request):
        qs = PersonaRRHH.objects.filter(is_deleted=True).select_related(
            "persona__tipo_documento",
            "persona__pais",
            "persona__departamento",
            "persona__ciudad",
        ).prefetch_related("especialidades")
        return Response(PersonaRRHHListSerializer(qs, many=True).data)

    @action(detail=False, methods=["get"], url_path="buscar")
    def buscar(self, request):
        nro_documento = request.query_params.get("nro_documento")
        if not nro_documento:
            return Response({"error": "nro_documento es requerido"}, status=400)
        try:
            persona = Persona.objects.select_related(
                "tipo_documento", "pais", "departamento", "ciudad"
            ).get(nro_documento=nro_documento, is_deleted=False)
            persona_data = PersonaListSerializer(persona).data
            try:
                prestador = PersonaRRHH.objects.select_related(
                    "persona__tipo_documento",
                    "persona__pais",
                    "persona__departamento",
                    "persona__ciudad",
                ).prefetch_related("especialidades").get(persona=persona, is_deleted=False)
                prestador_data = PersonaRRHHListSerializer(prestador).data
                es_prestador   = True
            except PersonaRRHH.DoesNotExist:
                prestador_data = None
                es_prestador   = False
            return Response({
                "persona":      persona_data,
                "personarrhh":  prestador_data,
                "es_prestador": es_prestador,
            })
        except Persona.DoesNotExist:
            return Response(
                {"persona": None, "personarrhh": None, "es_prestador": False},
                status=200,
            )

    @action(detail=False, methods=["get"], url_path="validar-matricula")
    def validar_matricula(self, request):
        nro        = request.query_params.get("nro_matricula", "").strip()
        exclude_id = request.query_params.get("exclude_id")
        if not nro:
            return Response({"disponible": True})
        qs = PersonaRRHH.objects.filter(nro_matricula=nro, is_deleted=False)
        if exclude_id:
            qs = qs.exclude(id=exclude_id)
        return Response({"disponible": not qs.exists()})

    @action(detail=False, methods=["get"], url_path="reporte-lista")
    def reporte_lista(self, request):
        prestadores = (
            PersonaRRHH.objects
            .filter(is_deleted=False)
            .select_related("persona")
            .prefetch_related("especialidades")
            .order_by("persona__razon_social")
        )
        CARGO_LABEL = {
            "medico": "Médico", "enfermero": "Enfermero/a",
            "administrativo": "Administrativo", "tecnico": "Técnico", "otro": "Otro",
        }
        filas = [
            {
                "nro":            i,
                "nombre":         p.persona.razon_social,
                "documento":      p.persona.nro_documento,
                "cargo":          CARGO_LABEL.get(p.cargo, p.cargo),
                "especialidades": ", ".join(e.descripcion for e in p.especialidades.all()) or "—",
                "matricula":      p.nro_matricula or "—",
                "estado":         p.get_estado_display(),
            }
            for i, p in enumerate(prestadores, start=1)
        ]
        hoy      = date.today()
        contexto = {"filas": filas, "fecha": hoy, "total": len(filas)}
        html     = render_to_string("informes/prestador_lista.html", contexto, request=request)
        try:
            import weasyprint
            pdf_bytes = weasyprint.HTML(
                string=html, base_url=request.build_absolute_uri("/")
            ).write_pdf()
        except Exception as e:
            return HttpResponse(f"Error generando PDF: {e}", status=500)
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="listado_prestadores.pdf"'
        return response

    @action(detail=False, methods=["get"], url_path="reporte-lista-excel")
    def reporte_lista_excel(self, request):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse("openpyxl no está instalado.", status=500)

        CARGO_LABEL = {
            "medico": "Médico", "enfermero": "Enfermero/a",
            "administrativo": "Administrativo", "tecnico": "Técnico", "otro": "Otro",
        }

        prestadores = (
            PersonaRRHH.objects
            .filter(is_deleted=False)
            .select_related("persona")
            .prefetch_related("especialidades")
            .order_by("persona__razon_social")
        )
        filas = [
            {
                "nro":            i,
                "nombre":         p.persona.razon_social,
                "documento":      p.persona.nro_documento,
                "cargo":          CARGO_LABEL.get(p.cargo, p.cargo),
                "especialidades": ", ".join(e.descripcion for e in p.especialidades.all()) or "—",
                "matricula":      p.nro_matricula or "—",
                "estado":         p.get_estado_display(),
            }
            for i, p in enumerate(prestadores, start=1)
        ]
        hoy = date.today()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listado de Prestadores"

        COLOR_PRIMARIO = "1A3A5C"
        COLOR_FILA_PAR = "F8FAFC"
        COLOR_BORDE    = "E8EDF2"

        fill_header  = PatternFill("solid", fgColor=COLOR_PRIMARIO)
        font_header  = Font(color="FFFFFF", bold=True, size=10)
        font_titulo  = Font(color=COLOR_PRIMARIO, bold=True, size=13)
        font_meta    = Font(color="555555", size=9)
        fill_par     = PatternFill("solid", fgColor=COLOR_FILA_PAR)
        thin_border  = Border(bottom=Side(style="thin", color=COLOR_BORDE))
        align_center = Alignment(horizontal="center", vertical="center")
        align_left   = Alignment(horizontal="left",   vertical="center")

        COLUMNAS    = 7
        fila_cursor = 1

        ws.merge_cells(f"A{fila_cursor}:{get_column_letter(COLUMNAS)}{fila_cursor}")
        celda = ws.cell(fila_cursor, 1, "Clínica Lichi — Listado de Prestadores")
        celda.font = font_titulo; celda.alignment = align_left
        ws.row_dimensions[fila_cursor].height = 20
        fila_cursor += 1

        ws.merge_cells(f"A{fila_cursor}:{get_column_letter(COLUMNAS)}{fila_cursor}")
        celda = ws.cell(fila_cursor, 1, f"Generado el {hoy.strftime('%d/%m/%Y')}  —  {len(filas)} registro{'s' if len(filas) != 1 else ''}")
        celda.font = font_meta; celda.alignment = align_left
        fila_cursor += 2

        for col, texto in enumerate(["N°", "Nombre completo", "Documento", "Cargo", "Especialidades", "Matrícula", "Estado"], start=1):
            c = ws.cell(fila_cursor, col, texto)
            c.fill = fill_header; c.font = font_header
            c.alignment = align_center if col == 1 else align_left
        ws.row_dimensions[fila_cursor].height = 18
        fila_cursor += 1

        for fila in filas:
            es_par  = (fila["nro"] % 2 == 0)
            valores = [fila["nro"], fila["nombre"], fila["documento"],
                       fila["cargo"], fila["especialidades"], fila["matricula"], fila["estado"]]
            for col, val in enumerate(valores, start=1):
                c = ws.cell(fila_cursor, col, val)
                if es_par: c.fill = fill_par
                c.border    = thin_border
                c.alignment = align_center if col == 1 else align_left
            ws.row_dimensions[fila_cursor].height = 15
            fila_cursor += 1

        for i, ancho in enumerate([6, 32, 16, 14, 36, 14, 12], start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="listado_prestadores_{hoy.strftime("%Y%m%d")}.xlsx"'
        return response
