from datetime import date

from django.http import HttpResponse
from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from .models import RegistroAuditoria
from .serializers import RegistroAuditoriaSerializer
from .utils import MODULOS


class RegistroAuditoriaViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class   = RegistroAuditoriaSerializer
    permission_classes = [IsAuthenticated]
    filter_backends    = [filters.SearchFilter, filters.OrderingFilter]
    search_fields      = ['tabla', 'accion', 'usuario__username']
    ordering_fields    = ['fecha', 'tabla', 'accion']
    ordering           = ['-fecha']

    def get_queryset(self):
        # Solo administradores pueden ver los registros de auditoría
        try:
            rol = self.request.user.perfil.rol
        except Exception:
            rol = None
        if rol != 'admin':
            raise PermissionDenied('Solo los administradores pueden acceder al registro de auditoría.')

        qs = RegistroAuditoria.objects.select_related('usuario').all()

        tabla   = self.request.query_params.get('tabla')
        accion  = self.request.query_params.get('accion')
        usuario = self.request.query_params.get('usuario')
        fecha_desde = self.request.query_params.get('fecha_desde')
        fecha_hasta = self.request.query_params.get('fecha_hasta')

        if tabla:
            qs = qs.filter(tabla__icontains=tabla)
        if accion:
            qs = qs.filter(accion=accion.upper())
        if usuario:
            qs = qs.filter(usuario__username__icontains=usuario)
        if fecha_desde:
            qs = qs.filter(fecha__date__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha__date__lte=fecha_hasta)

        modulo = self.request.query_params.get('modulo')
        if modulo:
            tablas = [k for k, v in MODULOS.items() if v == modulo]
            qs = qs.filter(tabla__in=tablas)

        return qs

    def _get_queryset_filtrado(self, request):
        qs = self.get_queryset()
        return qs

    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse('openpyxl no está instalado.', status=500)

        qs = self.get_queryset().select_related('usuario')
        hoy = date.today()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Registro de Auditoría'

        COLOR_PRIMARIO = '1A3A5C'
        COLOR_FILA_PAR = 'F8FAFC'
        COLOR_BORDE    = 'E8EDF2'

        fill_header  = PatternFill('solid', fgColor=COLOR_PRIMARIO)
        font_header  = Font(color='FFFFFF', bold=True, size=10)
        font_titulo  = Font(color=COLOR_PRIMARIO, bold=True, size=13)
        font_meta    = Font(color='555555', size=9)
        fill_par     = PatternFill('solid', fgColor=COLOR_FILA_PAR)
        thin_border  = Border(bottom=Side(style='thin', color=COLOR_BORDE))
        align_center = Alignment(horizontal='center', vertical='center')
        align_left   = Alignment(horizontal='left',   vertical='center')

        COLUMNAS    = 7
        fila_cursor = 1

        ws.merge_cells(f'A{fila_cursor}:{get_column_letter(COLUMNAS)}{fila_cursor}')
        c = ws.cell(fila_cursor, 1, 'Clínica Lichi — Registro de Auditoría')
        c.font = font_titulo; c.alignment = align_left
        ws.row_dimensions[fila_cursor].height = 20
        fila_cursor += 1

        total = qs.count()
        ws.merge_cells(f'A{fila_cursor}:{get_column_letter(COLUMNAS)}{fila_cursor}')
        c = ws.cell(fila_cursor, 1,
                    f"Generado el {hoy.strftime('%d/%m/%Y')}  —  "
                    f"{total} registro{'s' if total != 1 else ''}")
        c.font = font_meta; c.alignment = align_left
        fila_cursor += 2

        encabezados = ['N°', 'Fecha', 'Usuario', 'Acción', 'Tabla', 'Registro ID', 'IP']
        for col, texto in enumerate(encabezados, start=1):
            c = ws.cell(fila_cursor, col, texto)
            c.fill = fill_header; c.font = font_header
            c.alignment = align_center if col in (1, 6) else align_left
        ws.row_dimensions[fila_cursor].height = 18
        fila_cursor += 1

        for nro, reg in enumerate(qs, start=1):
            es_par = (nro % 2 == 0)
            fecha_fmt = reg.fecha.strftime('%d/%m/%Y %H:%M') if reg.fecha else ''
            usuario_str = reg.usuario.username if reg.usuario else '—'
            valores = [nro, fecha_fmt, usuario_str, reg.accion,
                       reg.tabla, reg.registro_id, reg.ip or '—']
            for col, val in enumerate(valores, start=1):
                c = ws.cell(fila_cursor, col, val)
                if es_par: c.fill = fill_par
                c.border    = thin_border
                c.alignment = align_center if col in (1, 6) else align_left
            ws.row_dimensions[fila_cursor].height = 15
            fila_cursor += 1

        for i, ancho in enumerate([6, 18, 20, 14, 24, 12, 16], start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="auditoria_{hoy.strftime("%Y%m%d")}.xlsx"'
        )
        return response